from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from login.models import Usuario, Unidad
from poa.models import Proyecto, MetaProyecto, Actividad, AvanceMensual, Evidencia, AuditoriaLog
from .decorators import auditor_required
import json
from decimal import Decimal
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook


from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


# ---
from django.db.models.functions import Coalesce
# --- NUEVAS IMPORTACIONES AÑADIDAS ---
from django.core.paginator import Paginator
from openpyxl.cell.cell import Cell
# ---

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from utils.exportacion import (
    generar_pdf_proyecto_detalle,
    generar_excel_proyecto_detalle,
    generar_pdf_unidades,
    generar_excel_unidades,
    generar_pdf_reporte_trimestral, 
    generar_excel_reporte_trimestral
)


@auditor_required
def dashboard_auditor(request):
    """Dashboard principal del auditor con estadísticas generales"""
    
    # Estadísticas generales (solo lectura)
    total_unidades = Usuario.objects.filter(rol='UNIDAD').count()
    total_proyectos = Proyecto.objects.count()
    proyectos_pendientes = Proyecto.objects.filter(estado='ENVIADO').count()
    proyectos_aprobados = Proyecto.objects.filter(estado='APROBADO').count()
    
    # Proyectos recientes
    proyectos_recientes = Proyecto.objects.select_related('unidad').order_by('-fecha_creacion')[:5]
    
    # Logs recientes de auditoría
    logs_recientes = AuditoriaLog.objects.select_related('usuario').order_by('-fecha')[:10]
    
    contexto = {
        'titulo': 'Panel de Auditoría',
        'total_unidades': total_unidades,
        'total_proyectos': total_proyectos,
        'proyectos_pendientes': proyectos_pendientes,
        'proyectos_aprobados': proyectos_aprobados,
        'proyectos_recientes': proyectos_recientes,
        'logs_recientes': logs_recientes,
    }
    
    return render(request, 'auditor/dashboard.html', contexto)


def _get_datos_trimestrales(busqueda_str=""):
    """
    Función de ayuda para calcular y filtrar los datos trimestrales.
    Usada por estadisticas_admin y las vistas de exportación.
    """
    unidades_con_datos = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    unidades_trimestrales_data = []

    for unidad in unidades_con_datos:
        nombre_unidad = unidad.unidad.nombre
        
        # Filtro de búsqueda
        if busqueda_str and busqueda_str.lower() not in nombre_unidad.lower():
            continue 
        
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        
        t1_avg, t2_avg, t3_avg, t4_avg = 0.0, 0.0, 0.0, 0.0
        
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            
            t1_data = avances.filter(mes__in=[1, 2, 3]).aggregate(avg=Avg('cumplimiento'))
            t1_avg = float(t1_data['avg'] or 0.0)
            
            t2_data = avances.filter(mes__in=[4, 5, 6]).aggregate(avg=Avg('cumplimiento'))
            t2_avg = float(t2_data['avg'] or 0.0)
            
            t3_data = avances.filter(mes__in=[7, 8, 9]).aggregate(avg=Avg('cumplimiento'))
            t3_avg = float(t3_data['avg'] or 0.0)
            
            t4_data = avances.filter(mes__in=[10, 11, 12]).aggregate(avg=Avg('cumplimiento'))
            t4_avg = float(t4_data['avg'] or 0.0)
            
        unidades_trimestrales_data.append({
            'nombre': nombre_unidad,
            't1': round(t1_avg, 2),
            't2': round(t2_avg, 2),
            't3': round(t3_avg, 2),
            't4': round(t4_avg, 2),
        })
        
    return unidades_trimestrales_data


@auditor_required
def estadisticas_auditor(request):
    """Página de estadísticas y gráficos (solo lectura)"""
    
    busqueda = request.GET.get('buscar', '')
    page_number = request.GET.get('page', 1)
    active_tab = request.GET.get('tab', 'general')
    
    unidades_con_datos = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    
    unidades_data = {
        'nombres': [],
        'cumplimiento': [],
        'proyectos': [],
        'metricas': []
    }
    
    for unidad in unidades_con_datos:
        unidades_data['nombres'].append(unidad.unidad.nombre)
        
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento_promedio = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg']
            
            if cumplimiento_promedio is not None:
                cumplimiento_promedio = float(cumplimiento_promedio)
            else:
                cumplimiento_promedio = 0.0
            
            unidades_data['cumplimiento'].append(round(cumplimiento_promedio, 2))
            
            total_evidencias = Evidencia.objects.filter(actividad__meta__proyecto=poa_aprobado).count()
            total_actividades = Actividad.objects.filter(meta__proyecto=poa_aprobado).count()
            unidades_data['metricas'].append([
                round(cumplimiento_promedio, 2),
                min(total_evidencias * 10, 100),
                min(total_actividades * 5, 100),
                85,
                90
            ])
        else:
            unidades_data['cumplimiento'].append(0.0)
            unidades_data['metricas'].append([0, 0, 0, 0, 0])
        
        total_proyectos_unidad = Proyecto.objects.filter(unidad=unidad).count()
        unidades_data['proyectos'].append(total_proyectos_unidad)
    
    proyectos_data = {
        'estados': [
            Proyecto.objects.filter(estado='APROBADO').count(),
            Proyecto.objects.filter(estado='ENVIADO').count(),
            Proyecto.objects.filter(estado='RECHAZADO').count(),
            Proyecto.objects.filter(estado='BORRADOR').count()
        ],
        'programado_mensual': [],
        'realizado_mensual': []
    }
    
    for mes in range(1, 13):
        programado = AvanceMensual.objects.filter(mes=mes).aggregate(
            total=Count('cantidad_programada_mes')
        )['total'] or 0
        realizado = AvanceMensual.objects.filter(mes=mes).aggregate(
            total=Count('cantidad_realizada')
        )['total'] or 0
        proyectos_data['programado_mensual'].append(programado)
        proyectos_data['realizado_mensual'].append(realizado)
    
    unidades_trimestrales_filtradas = _get_datos_trimestrales(busqueda)
    paginator = Paginator(unidades_trimestrales_filtradas, 6) 
    page_obj = paginator.get_page(page_number)
    
    total_unidades = unidades_con_datos.count()
    total_proyectos = Proyecto.objects.count()
    proyectos_aprobados = Proyecto.objects.filter(estado='APROBADO').count()
    proyectos_pendientes = Proyecto.objects.filter(estado='ENVIADO').count()
    
    contexto = {
        'titulo': 'Estadísticas del Sistema',
        'unidades_data': json.dumps(unidades_data),
        'proyectos_data': json.dumps(proyectos_data),
        'total_unidades': total_unidades,
        'total_proyectos': total_proyectos,
        'proyectos_aprobados': proyectos_aprobados,
        'proyectos_pendientes': proyectos_pendientes,
        
        'unidades_trimestrales_page': page_obj,
        'busqueda': busqueda,
        'active_tab': active_tab,
    }
    
    return render(request, 'auditor/estadisticas.html', contexto)


@auditor_required
def ver_usuarios(request):
    """Lista todos los usuarios del sistema (solo lectura)"""
    
    busqueda = request.GET.get('buscar', '')
    
    usuarios = Usuario.objects.select_related('unidad').all()
    
    if busqueda:
        usuarios = usuarios.filter(
            Q(email__icontains=busqueda) |
            Q(unidad__nombre__icontains=busqueda) |
            Q(rol__icontains=busqueda)
        )
    
    usuarios = usuarios.order_by('rol', 'email')
    
    contexto = {
        'titulo': 'Usuarios del Sistema',
        'usuarios': usuarios,
        'busqueda': busqueda,
    }
    
    return render(request, 'auditor/ver_usuarios.html', contexto)


@auditor_required
def ver_logs(request):
    
    busqueda = request.GET.get('buscar', '')
    accion_filtro = request.GET.get('accion', '')
    
    logs = AuditoriaLog.objects.select_related('usuario').all()
    
    if busqueda:
        logs = logs.filter(
            Q(usuario__email__icontains=busqueda) |
            Q(tabla__icontains=busqueda) |
            Q(accion__icontains=busqueda)
        )
    
    if accion_filtro:
        logs = logs.filter(accion=accion_filtro)
    
    logs = logs.order_by('-fecha')[:100] 
    
    acciones = AuditoriaLog.objects.values_list('accion', flat=True).distinct()
    
    contexto = {
        'titulo': 'Logs de Auditoría',
        'logs': logs,
        'busqueda': busqueda,
        'accion_filtro': accion_filtro,
        'acciones': acciones,
    }
    
    return render(request, 'auditor/ver_logs.html', contexto)


@auditor_required
def ver_proyectos(request):
    """Lista todos los proyectos de todas las unidades (solo lectura)"""
    
    busqueda = request.GET.get('buscar', '')
    estado_filtro = request.GET.get('estado', '')
    
    proyectos = Proyecto.objects.select_related('unidad').all()
    
    if busqueda:
        proyectos = proyectos.filter(
            Q(nombre__icontains=busqueda) |
            Q(unidad__nombre__icontains=busqueda)
        )
    
    if estado_filtro:
        proyectos = proyectos.filter(estado=estado_filtro)
    
    proyectos = proyectos.order_by('-fecha_creacion')
    
    contexto = {
        'titulo': 'Todos los Proyectos',
        'proyectos': proyectos,
        'busqueda': busqueda,
        'estado_filtro': estado_filtro,
    }
    
    return render(request, 'auditor/ver_proyectos.html', contexto)


@auditor_required
def detalle_proyecto_auditor(request, proyecto_id):
    """Muestra el detalle completo de un proyecto (solo lectura)"""
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    metas = proyecto.metas.prefetch_related('actividades__avances', 'actividades__evidencias')
    
    total_actividades = Actividad.objects.filter(meta__proyecto=proyecto).count()
    
    presupuesto_total = Actividad.objects.filter(meta__proyecto=proyecto).aggregate(
        total=Sum('total_recursos')
    )['total'] or Decimal('0.00')
    
    total_evidencias = Evidencia.objects.filter(actividad__meta__proyecto=proyecto).count()
    
    for meta in metas:
        for actividad in meta.actividades.all():
            for avance in actividad.avances.all():
                avance.evidencias_count = Evidencia.objects.filter(
                    actividad=actividad,
                    mes=avance.mes
                ).count()
    
    contexto = {
        'titulo': f'Detalle: {proyecto.nombre}',
        'proyecto': proyecto,
        'metas': metas,
        'total_actividades': total_actividades,
        'presupuesto_total': presupuesto_total,
        'total_evidencias': total_evidencias,
    }
    
    return render(request, 'auditor/detalle_proyecto.html', contexto)


@auditor_required
def ver_unidades(request):
    """Lista todas las unidades con sus estadísticas (solo lectura)"""
    
    busqueda = request.GET.get('buscar', '')
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    
    if busqueda:
        unidades = unidades.filter(
            Q(unidad__nombre__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    unidades = unidades.annotate(
        total_proyectos=Count('proyectos'),
        count_proyectos_aprobados=Count('proyectos', filter=Q(proyectos__estado='APROBADO'))
    )
    
    unidades_con_rendimiento = []
    for unidad in unidades:
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento_promedio = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            unidad.rendimiento = round(cumplimiento_promedio, 2)
        else:
            unidad.rendimiento = 0
        unidades_con_rendimiento.append(unidad)
    
    contexto = {
        'titulo': 'Unidades del Sistema',
        'unidades': unidades_con_rendimiento,
        'busqueda': busqueda,
    }
    
    return render(request, 'auditor/ver_unidades.html', contexto)


@auditor_required
def exportar_estadisticas_pdf(request):
    """Exporta las estadísticas del sistema a PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="estadisticas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []
    styles = getSampleStyleSheet()
    
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elementos.append(Paragraph('Estadísticas del Sistema POA', titulo_style))
    elementos.append(Paragraph(f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elementos.append(Spacer(1, 0.5*inch))
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    
    data = [['Unidad', 'Proyectos', 'Aprobados', 'Cumplimiento %']]
    
    for unidad in unidades:
        total_proyectos = Proyecto.objects.filter(unidad=unidad).count()
        proyectos_aprobados = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').count()
        
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            cumplimiento = round(cumplimiento, 2)
        else:
            cumplimiento = 0
        
        data.append([
            unidad.unidad.nombre,
            str(total_proyectos),
            str(proyectos_aprobados),
            f'{cumplimiento}%'
        ])
    
    tabla = Table(data, colWidths=[3*inch, 1*inch, 1*inch, 1.5*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c4a6e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elementos.append(tabla)
    doc.build(elementos)
    
    return response


@auditor_required
def exportar_estadisticas_excel(request):
    """Exporta las estadísticas del sistema a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas"
    
    header_fill = PatternFill(start_color="0c4a6e", end_color="0c4a6e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Estadísticas del Sistema POA'
    ws['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['Unidad', 'Proyectos', 'Aprobados', 'Cumplimiento %']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    row = 5
    
    for unidad in unidades:
        total_proyectos = Proyecto.objects.filter(unidad=unidad).count()
        proyectos_aprobados = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').count()
        
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            cumplimiento = round(cumplimiento, 2)
        else:
            cumplimiento = 0
        
        ws.cell(row=row, column=1, value=unidad.unidad.nombre).border = border
        ws.cell(row=row, column=2, value=total_proyectos).border = border
        ws.cell(row=row, column=3, value=proyectos_aprobados).border = border
        ws.cell(row=row, column=4, value=f'{cumplimiento}%').border = border
        
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="estadisticas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@auditor_required
def exportar_proyecto_detalle_pdf(request, proyecto_id):
    """Exporta el detalle COMPLETO de un proyecto a PDF - Usa función compartida"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    response = generar_pdf_proyecto_detalle(proyecto, request.user)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_PDF_AUDITOR',
        tabla='Proyecto',
        registro_id=proyecto.id,
        datos_nuevos={'tipo': 'PDF_DETALLADO', 'proyecto': proyecto.nombre},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@auditor_required
def exportar_proyecto_detalle_excel(request, proyecto_id):
    """Exporta el detalle COMPLETO de un proyecto a Excel - Usa función compartida"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    response = generar_excel_proyecto_detalle(proyecto, request.user)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_EXCEL_AUDITOR',
        tabla='Proyecto',
        registro_id=proyecto.id,
        datos_nuevos={'tipo': 'EXCEL_DETALLADO', 'proyecto': proyecto.nombre},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@auditor_required
def exportar_unidades_pdf(request):
    """Exporta el reporte de todas las unidades con su cumplimiento a PDF - Usa función compartida"""
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    unidades = unidades.annotate(
        total_proyectos=Count('proyectos'),
        count_proyectos_aprobados=Count('proyectos', filter=Q(proyectos__estado='APROBADO'))
    )
    
    unidades_con_rendimiento = []
    for unidad in unidades:
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento_promedio = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            unidad.rendimiento = round(cumplimiento_promedio, 2)
        else:
            unidad.rendimiento = 0
        
        if unidad.rendimiento >= 80:
            unidad.categoria = 'Excelente'
        elif unidad.rendimiento >= 60:
            unidad.categoria = 'Bueno'
        elif unidad.rendimiento >= 40:
            unidad.categoria = 'Regular'
        else:
            unidad.categoria = 'Bajo'
        
        unidades_con_rendimiento.append(unidad)
    
    response = generar_pdf_unidades(unidades_con_rendimiento, request.user)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_PDF_AUDITOR',
        tabla='Usuario',
        registro_id=0,
        datos_nuevos={'tipo': 'PDF_UNIDADES', 'total_unidades': len(unidades_con_rendimiento)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@auditor_required
def exportar_unidades_excel(request):
    """Exporta el reporte de todas las unidades con su cumplimiento a Excel - Usa función compartida"""
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    unidades = unidades.annotate(
        total_proyectos=Count('proyectos'),
        count_proyectos_aprobados=Count('proyectos', filter=Q(proyectos__estado='APROBADO'))
    )
    
    unidades_con_rendimiento = []
    for unidad in unidades:
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento_promedio = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            unidad.rendimiento = round(cumplimiento_promedio, 2)
        else:
            unidad.rendimiento = 0
        
        if unidad.rendimiento >= 80:
            unidad.categoria = 'Excelente'
            unidad.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        elif unidad.rendimiento >= 60:
            unidad.categoria = 'Bueno'
            unidad.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
        elif unidad.rendimiento >= 40:
            unidad.categoria = 'Regular'
            unidad.fill = PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid")
        else:
            unidad.categoria = 'Bajo'
            unidad.fill = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")
        
        unidades_con_rendimiento.append(unidad)
    
    response = generar_excel_unidades(unidades_con_rendimiento, request.user)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_EXCEL_AUDITOR',
        tabla='Usuario',
        registro_id=0,
        datos_nuevos={'tipo': 'EXCEL_UNIDADES', 'total_unidades': len(unidades_con_rendimiento)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@auditor_required
def exportar_proyectos_pdf(request):
    """
    Exporta una lista detallada de proyectos a PDF, incluyendo
    estadísticas de metas, actividades y presupuesto.
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Proyectos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, 
                            topMargin=0.5*inch, bottomMargin=0.5*inch, 
                            leftMargin=0.5*inch, rightMargin=0.5*inch)
    elementos = []
    styles = getSampleStyleSheet()
    
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20, 
        textColor=colors.HexColor('#d97706'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    cell_style = ParagraphStyle(
        'BodyCell',
        parent=styles['Normal'],
        fontSize=8, 
        leading=10
    )
    
    elementos.append(Paragraph('Reporte Detallado de Proyectos', titulo_style))
    
    fecha_gen_local = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
    elementos.append(Paragraph(f'Generado: {fecha_gen_local}', styles['Normal']))
    elementos.append(Spacer(1, 0.3*inch))
    
    proyectos = Proyecto.objects.select_related(
        'unidad__unidad'
    ).annotate(
        count_metas=Count('metas', distinct=True),
        count_actividades=Count('metas__actividades', distinct=True),
        sum_presupuesto=Coalesce(
            Sum('metas__actividades__total_recursos'),
            Decimal('0.00')
        )
    ).order_by('unidad__unidad__nombre', '-anio', 'nombre').all()
    
    data = [
        ['Unidad', 'Proyecto', 'Año', 'Estado', 'Metas', 'Activ.', 'Presupuesto', 'Fecha Creac.']
    ]
    
    for proyecto in proyectos:
        unidad_p = Paragraph(proyecto.unidad.unidad.nombre, cell_style)
        proyecto_p = Paragraph(proyecto.nombre, cell_style)
        
        fecha_creacion_local = timezone.localtime(proyecto.fecha_creacion)
        fecha_str = fecha_creacion_local.strftime('%d/%m/%Y')
        
        presup_str = f'${proyecto.sum_presupuesto:,.2f}'

        data.append([
            unidad_p,
            proyecto_p,
            proyecto.anio,
            proyecto.get_estado_display(), 
            proyecto.count_metas,
            proyecto.count_actividades,
            presup_str,
            fecha_str
        ])
    
    # --- Definición de Columnas y Estilos de Tabla ---
    colWidths = [
        1.3*inch,  # Unidad
        1.8*inch,  # Proyecto
        0.4*inch,  # Año
        0.8*inch,  # Estado
        0.5*inch,  # Metas
        0.5*inch,  # Activ.
        1.0*inch,  # Presupuesto
        0.9*inch   # Fecha Creac.
    ]
    
    tabla = Table(data, colWidths=colWidths)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d97706')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'), 
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F7F7F7')), 
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D0D0D0')),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    elementos.append(tabla)
    
    doc.build(elementos)
    
    return response

@auditor_required
def exportar_proyectos_excel(request):
    """
    Exporta una lista detallada de proyectos, incluyendo estadísticas
    de metas, actividades y presupuesto.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    
    # Estilos 
    header_fill = PatternFill(start_color="d97706", end_color="d97706", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- Encabezado ---
    ws.merge_cells('A1:K1')
    ws['A1'] = 'Reporte Detallado de Proyectos'
    ws['A1'].font = Font(bold=True, size=16, color="d97706")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:K2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # --- Headers) ---
    headers = [
        'Unidad', 'Proyecto', 'Año', 'Estado', 'Total Metas', 
        'Total Actividades', 'Presupuesto Total', 'Total Evidencias', 
        'Aprobado por', 'Fecha Creación', 'Fecha Aprobación'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    proyectos = Proyecto.objects.select_related(
        'unidad__unidad', 'aprobado_por'
    ).annotate(
        count_metas=Count('metas', distinct=True),
        count_actividades=Count('metas__actividades', distinct=True),
        sum_presupuesto=Coalesce(
            Sum('metas__actividades__total_recursos'),
            Decimal('0.00')
        ),
        count_evidencias=Count('metas__actividades__evidencias', distinct=True)
    ).order_by('unidad__unidad__nombre', '-anio', 'nombre')
    
    row = 5
    
    for proyecto in proyectos:
        ws.cell(row=row, column=1, value=proyecto.unidad.unidad.nombre).border = border
        ws.cell(row=row, column=2, value=proyecto.nombre).border = border
        ws.cell(row=row, column=3, value=proyecto.anio).border = border
        ws.cell(row=row, column=4, value=proyecto.get_estado_display()).border = border 
        
        ws.cell(row=row, column=5, value=proyecto.count_metas).border = border
        ws.cell(row=row, column=6, value=proyecto.count_actividades).border = border
        ws.cell(row=row, column=7, value=proyecto.sum_presupuesto).border = border
        ws.cell(row=row, column=7).number_format = '"$"#,##0.00' # Formato de moneda
        ws.cell(row=row, column=8, value=proyecto.count_evidencias).border = border
        
        ws.cell(row=row, column=9, value=proyecto.aprobado_por.email if proyecto.aprobado_por else 'N/A').border = border
        
        fecha_creacion_local = timezone.localtime(proyecto.fecha_creacion)
        ws.cell(row=row, column=10, value=fecha_creacion_local.replace(tzinfo=None)).border = border
        ws.cell(row=row, column=10).number_format = 'dd/mm/yyyy hh:mm'

        # Convertir fecha_aprobacion (que puede ser Nula)
        if proyecto.fecha_aprobacion:
            fecha_aprobacion_local = timezone.localtime(proyecto.fecha_aprobacion)
            ws.cell(row=row, column=11, value=fecha_aprobacion_local.replace(tzinfo=None)).border = border
            ws.cell(row=row, column=11).number_format = 'dd/mm/yyyy hh:mm'
        else:
            ws.cell(row=row, column=11, value='N/A').border = border    
        
        row += 1
    
    # --- Ajuste de anchos de columna ---
    ws.column_dimensions['A'].width = 30  # Unidad
    ws.column_dimensions['B'].width = 45  # Proyecto
    ws.column_dimensions['C'].width = 8   # Año
    ws.column_dimensions['D'].width = 15  # Estado
    ws.column_dimensions['E'].width = 12  # Total Metas
    ws.column_dimensions['F'].width = 12  # Total Actividades
    ws.column_dimensions['G'].width = 18  # Presupuesto Total
    ws.column_dimensions['H'].width = 12  # Total Evidencias
    ws.column_dimensions['I'].width = 25  # Aprobado por
    ws.column_dimensions['J'].width = 18  # Fecha Creación
    ws.column_dimensions['K'].width = 18  # Fecha Aprobación
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Proyectos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@auditor_required
def exportar_usuarios_pdf(request):
    """Exporta la lista de usuarios a PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []
    styles = getSampleStyleSheet()
    
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#d97706'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elementos.append(Paragraph('Usuarios del Sistema', titulo_style))
    elementos.append(Paragraph(f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elementos.append(Spacer(1, 0.5*inch))
    
    usuarios = Usuario.objects.select_related('unidad').all()
    
    data = [['Email', 'Rol', 'Unidad', 'Activo']]
    
    for usuario in usuarios:
        unidad_nombre = usuario.unidad.nombre if usuario.unidad else 'N/A'
        data.append([
            usuario.email[:35],
            usuario.rol,
            unidad_nombre[:30],
            'Sí' if usuario.is_active else 'No'
        ])
    
    tabla = Table(data, colWidths=[2.5*inch, 1.5*inch, 2*inch, 1*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d97706')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elementos.append(tabla)
    doc.build(elementos)
    
    return response


@auditor_required
def exportar_usuarios_excel(request):
    """Exporta la lista de usuarios a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    header_fill = PatternFill(start_color="d97706", end_color="d97706", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Usuarios del Sistema'
    ws['A1'].font = Font(bold=True, size=16, color="d97706")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['Email', 'Rol', 'Unidad', 'Activo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    usuarios = Usuario.objects.select_related('unidad').all()
    row = 5
    
    for usuario in usuarios:
        unidad_nombre = usuario.unidad.nombre if usuario.unidad else 'N/A'
        ws.cell(row=row, column=1, value=usuario.email).border = border
        ws.cell(row=row, column=2, value=usuario.rol).border = border
        ws.cell(row=row, column=3, value=unidad_nombre).border = border
        ws.cell(row=row, column=4, value='Sí' if usuario.is_active else 'No').border = border
        
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@auditor_required
def exportar_logs_pdf(request):
    """Exporta los logs de auditoría a PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []
    styles = getSampleStyleSheet()
    
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#d97706'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elementos.append(Paragraph('Logs de Auditoría', titulo_style))
    elementos.append(Paragraph(f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elementos.append(Spacer(1, 0.5*inch))
    
    logs = AuditoriaLog.objects.select_related('usuario').order_by('-fecha')[:100]
    
    data = [['Fecha', 'Usuario', 'Acción', 'Tabla']]
    
    for log in logs:
        usuario_email = log.usuario.email if log.usuario else 'Sistema'
        data.append([
            log.fecha.strftime('%d/%m/%Y %H:%M'),
            usuario_email[:25],
            log.accion,
            log.tabla[:20]
        ])
    
    tabla = Table(data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d97706')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elementos.append(tabla)
    doc.build(elementos)
    
    return response


@auditor_required
def exportar_logs_excel(request):
    """Exporta los logs de auditoría a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    
    header_fill = PatternFill(start_color="d97706", end_color="d97706", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Logs de Auditoría'
    ws['A1'].font = Font(bold=True, size=16, color="d97706")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['Fecha', 'Usuario', 'Acción', 'Tabla', ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    logs = AuditoriaLog.objects.select_related('usuario').order_by('-fecha')[:500]
    row = 5
    
    for log in logs:
        usuario_email = log.usuario.email if log.usuario else 'Sistema'
        ws.cell(row=row, column=1, value=log.fecha.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row, column=2, value=usuario_email).border = border
        ws.cell(row=row, column=3, value=log.accion).border = border
        ws.cell(row=row, column=4, value=log.tabla).border = border
        
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response



@auditor_required
def exportar_reporte_trimestral_pdf(request):
    """Exporta el reporte trimestral (filtrado) a PDF para Auditor"""
    
    busqueda = request.GET.get('buscar', '')
    
    unidades_trimestrales = _get_datos_trimestrales(busqueda)
    
    response = generar_pdf_reporte_trimestral(unidades_trimestrales, request.user, busqueda)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_PDF_AUDITOR',
        tabla='Reporte',
        registro_id=0,
        datos_nuevos={'tipo': 'PDF_REPORTE_TRIMESTRAL', 'filtro': busqueda, 'total': len(unidades_trimestrales)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response

@auditor_required
def exportar_reporte_trimestral_excel(request):
    """Exporta el reporte trimestral (filtrado) a Excel para Auditor"""
    
    busqueda = request.GET.get('buscar', '')
    
    unidades_trimestrales = _get_datos_trimestrales(busqueda)
    
    response = generar_excel_reporte_trimestral(unidades_trimestrales, request.user, busqueda)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_EXCEL_AUDITOR',
        tabla='Reporte',
        registro_id=0,
        datos_nuevos={'tipo': 'EXCEL_REPORTE_TRIMESTRAL', 'filtro': busqueda, 'total': len(unidades_trimestrales)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response