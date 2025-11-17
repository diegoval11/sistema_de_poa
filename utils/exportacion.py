"""
Módulo de utilidades compartidas para exportación de reportes
Usado por administrador y auditor para evitar duplicación de código
"""
from django.http import HttpResponse
from django.db.models import Sum, Avg, Count, Q
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
from decimal import Decimal

from poa.models import Actividad, AvanceMensual, Evidencia


def generar_pdf_proyecto_detalle(proyecto, usuario):
    """
    Genera un PDF detallado de un proyecto con toda su información
    Usado por administrador y auditor
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="POA_Detallado_{proyecto.unidad.unidad.nombre}_{proyecto.anio}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.4*inch, bottomMargin=0.4*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elementos = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=15,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitulo_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Heading3'],
        fontSize=11,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    
    small_style = ParagraphStyle(
        'SmallText',
        parent=styles['Normal'],
        fontSize=7,
        leading=9
    )
    
    # Encabezado del documento
    elementos.append(Paragraph('PLAN OPERATIVO ANUAL (POA)', titulo_style))
    elementos.append(Paragraph(f'<b>Proyecto:</b> {proyecto.nombre}', styles['Normal']))
    elementos.append(Paragraph(f'<b>Unidad Responsable:</b> {proyecto.unidad.unidad.nombre}', styles['Normal']))
    elementos.append(Paragraph(f'<b>Año de Ejecución:</b> {proyecto.anio}', styles['Normal']))
    elementos.append(Paragraph(f'<b>Estado:</b> {proyecto.get_estado_display()}', styles['Normal']))
    elementos.append(Paragraph(f'<b>Fecha de Generación:</b> {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elementos.append(Spacer(1, 0.3*inch))
    
    # Información general
    elementos.append(Paragraph('INFORMACIÓN GENERAL', subtitulo_style))
    
    info_data = [
        ['Objetivo de la Unidad:', Paragraph(proyecto.objetivo_unidad or 'No especificado', styles['Normal'])],
        ['Fecha de Creación:', proyecto.fecha_creacion.strftime('%d/%m/%Y %H:%M')],
        ['Última Modificación:', proyecto.fecha_modificacion.strftime('%d/%m/%Y %H:%M')],
    ]
    
    if proyecto.estado == 'APROBADO':
        info_data.append(['Aprobado por:', proyecto.aprobado_por.email if proyecto.aprobado_por else 'N/A'])
        info_data.append(['Fecha de Aprobación:', proyecto.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if proyecto.fecha_aprobacion else 'N/A'])
    elif proyecto.estado == 'RECHAZADO':
        info_data.append(['Motivo de Rechazo:', Paragraph(proyecto.motivo_rechazo or 'No especificado', styles['Normal'])])
    
    info_tabla = Table(info_data, colWidths=[2*inch, 4.5*inch])
    info_tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(info_tabla)
    elementos.append(Spacer(1, 0.3*inch))
    
    # Resumen de estadísticas
    total_actividades = Actividad.objects.filter(meta__proyecto=proyecto).count()
    presupuesto_total = Actividad.objects.filter(meta__proyecto=proyecto).aggregate(
        total=Sum('total_recursos')
    )['total'] or Decimal('0.00')
    total_evidencias = Evidencia.objects.filter(actividad__meta__proyecto=proyecto).count()
    
    elementos.append(Paragraph('RESUMEN EJECUTIVO', subtitulo_style))
    
    resumen_data = [
        ['Total de Metas:', str(proyecto.metas.count())],
        ['Total de Actividades:', str(total_actividades)],
        ['Presupuesto Total:', f'${presupuesto_total:,.2f}'],
        ['Total de Evidencias:', str(total_evidencias)],
    ]
    
    resumen_tabla = Table(resumen_data, colWidths=[2*inch, 4.5*inch])
    resumen_tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dbeafe')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(resumen_tabla)
    elementos.append(Spacer(1, 0.3*inch))
    
    # Metas y actividades detalladas
    elementos.append(Paragraph('METAS Y ACTIVIDADES DETALLADAS', subtitulo_style))
    elementos.append(Spacer(1, 0.15*inch))
    
    metas = proyecto.metas.prefetch_related('actividades__avances', 'actividades__evidencias').all()
    
    for idx_meta, meta in enumerate(metas, 1):
        # Título de la meta
        elementos.append(Paragraph(f'META {idx_meta}: {meta.descripcion}', meta_style))
        
        actividades_data = [['#', 'Actividad', 'U.M.', 'Cant.', 'Recursos']]
        
        for idx_act, actividad in enumerate(meta.actividades.all(), 1):
            actividades_data.append([
                str(idx_act),
                Paragraph(actividad.descripcion, small_style),
                Paragraph(actividad.unidad_medida[:12], small_style),
                str(actividad.cantidad_programada),
                f'${actividad.total_recursos:,.0f}'
            ])
        
        if len(actividades_data) > 1:
            act_tabla = Table(actividades_data, colWidths=[0.3*inch, 3.5*inch, 0.7*inch, 0.5*inch, 1*inch])
            act_tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c4a6e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))
            elementos.append(act_tabla)
            elementos.append(Spacer(1, 0.12*inch))
            
            # Programación mensual para cada actividad
            for actividad in meta.actividades.all():
                if actividad.avances.exists():
                    elementos.append(Paragraph(f'Programación: {actividad.descripcion[:60]}...', styles['Heading4']))
                    
                    prog_data = [['Mes', 'Prog.', 'Real.', '%', 'Estado']]
                    
                    for avance in actividad.avances.all().order_by('mes'):
                        estado = 'Excelente' if avance.cumplimiento >= 90 else 'Bueno' if avance.cumplimiento >= 70 else 'Regular' if avance.cumplimiento >= 50 else 'Deficiente'
                        prog_data.append([
                            avance.get_mes_display()[:3],
                            str(avance.cantidad_programada_mes),
                            str(avance.cantidad_realizada),
                            f'{avance.cumplimiento:.0f}%',
                            estado
                        ])
                    
                    prog_tabla = Table(prog_data, colWidths=[0.9*inch, 0.7*inch, 0.7*inch, 0.7*inch, 1*inch])
                    prog_tabla.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 7),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                        ('TOPPADDING', (0, 0), (-1, 0), 5),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eff6ff')),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('FONTSIZE', (0, 1), (-1, -1), 6),
                        ('TOPPADDING', (0, 1), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                    ]))
                    elementos.append(prog_tabla)
                    elementos.append(Spacer(1, 0.08*inch))
                    
                    evidencias = actividad.evidencias.all()
                    if evidencias.exists():
                        elementos.append(Paragraph(f'Evidencias ({evidencias.count()})', styles['Heading4']))
                        
                        evid_data = [['Tipo', 'Mes', 'Descripción', 'Fecha']]
                        
                        for evidencia in evidencias[:8]:
                            evid_data.append([
                                evidencia.tipo[:10],
                                evidencia.get_mes_display()[:3] if evidencia.mes else 'N/A',
                                Paragraph(evidencia.descripcion[:60] if evidencia.descripcion else 'Sin descripción', small_style),
                                evidencia.fecha_subida.strftime('%d/%m/%y')
                            ])
                        
                        evid_tabla = Table(evid_data, colWidths=[0.8*inch, 0.7*inch, 2.8*inch, 0.8*inch])
                        evid_tabla.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 7),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                            ('TOPPADDING', (0, 0), (-1, 0), 5),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf4')),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('FONTSIZE', (0, 1), (-1, -1), 6),
                            ('LEFTPADDING', (0, 0), (-1, -1), 3),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                            ('TOPPADDING', (0, 1), (-1, -1), 3),
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                        ]))
                        elementos.append(evid_tabla)
                        elementos.append(Spacer(1, 0.1*inch))
        
        elementos.append(PageBreak())
    
    # Pie de página
    elementos.append(Spacer(1, 0.5*inch))
    elementos.append(Paragraph('_' * 80, styles['Normal']))
    elementos.append(Paragraph(f'Documento generado automáticamente el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}', styles['Normal']))
    elementos.append(Paragraph('Sistema de Gestión POA - Alcaldía', styles['Normal']))
    
    doc.build(elementos)
    return response


def generar_excel_proyecto_detalle(proyecto, usuario):
    """
    Genera un Excel detallado de un proyecto con toda su información
    Usado por administrador y auditor
    """
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="0c4a6e", end_color="0c4a6e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Hoja 1: Información General
    ws_info = wb.active
    ws_info.title = "Información General"
    
    ws_info.merge_cells('A1:E1')
    ws_info['A1'] = 'PLAN OPERATIVO ANUAL (POA)'
    ws_info['A1'].font = Font(bold=True, size=18, color="0c4a6e")
    ws_info['A1'].alignment = Alignment(horizontal='center')
    
    ws_info.merge_cells('A2:E2')
    ws_info['A2'] = f'Proyecto: {proyecto.nombre}'
    ws_info['A2'].font = Font(bold=True, size=14)
    ws_info['A2'].alignment = Alignment(horizontal='center')
    
    ws_info.merge_cells('A3:E3')
    ws_info['A3'] = f'Unidad: {proyecto.unidad.unidad.nombre} - Año: {proyecto.anio}'
    ws_info['A3'].alignment = Alignment(horizontal='center')
    
    row = 5
    ws_info.cell(row=row, column=1, value='Campo').fill = header_fill
    ws_info.cell(row=row, column=1).font = header_font
    ws_info.cell(row=row, column=2, value='Valor').fill = header_fill
    ws_info.cell(row=row, column=2).font = header_font
    
    row += 1
    info_fields = [
        ('Objetivo de la Unidad', proyecto.objetivo_unidad or 'No especificado'),
        ('Estado', proyecto.get_estado_display()),
        ('Fecha de Creación', proyecto.fecha_creacion.strftime('%d/%m/%Y %H:%M')),
        ('Última Modificación', proyecto.fecha_modificacion.strftime('%d/%m/%Y %H:%M')),
    ]
    
    if proyecto.estado == 'APROBADO':
        info_fields.append(('Aprobado por', proyecto.aprobado_por.email if proyecto.aprobado_por else 'N/A'))
        info_fields.append(('Fecha de Aprobación', proyecto.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if proyecto.fecha_aprobacion else 'N/A'))
    elif proyecto.estado == 'RECHAZADO':
        info_fields.append(('Motivo de Rechazo', proyecto.motivo_rechazo or 'No especificado'))
    
    for field, value in info_fields:
        ws_info.cell(row=row, column=1, value=field).border = border
        ws_info.cell(row=row, column=1).font = Font(bold=True)
        ws_info.cell(row=row, column=2, value=value).border = border
        row += 1
    
    ws_info.column_dimensions['A'].width = 30
    ws_info.column_dimensions['B'].width = 60
    
    # Hoja 2: Resumen Ejecutivo
    ws_resumen = wb.create_sheet("Resumen Ejecutivo")
    
    total_actividades = Actividad.objects.filter(meta__proyecto=proyecto).count()
    presupuesto_total = Actividad.objects.filter(meta__proyecto=proyecto).aggregate(
        total=Sum('total_recursos')
    )['total'] or Decimal('0.00')
    total_evidencias = Evidencia.objects.filter(actividad__meta__proyecto=proyecto).count()
    
    ws_resumen.merge_cells('A1:B1')
    ws_resumen['A1'] = 'RESUMEN EJECUTIVO'
    ws_resumen['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    resumen_data = [
        ('Total de Metas', proyecto.metas.count()),
        ('Total de Actividades', total_actividades),
        ('Presupuesto Total', f'${presupuesto_total:,.2f}'),
        ('Total de Evidencias', total_evidencias),
    ]
    
    for field, value in resumen_data:
        ws_resumen.cell(row=row, column=1, value=field).font = Font(bold=True)
        ws_resumen.cell(row=row, column=1).border = border
        ws_resumen.cell(row=row, column=2, value=value).border = border
        row += 1
    
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 30
    
    # Hoja 3: Metas y Actividades
    ws_metas = wb.create_sheet("Metas y Actividades")
    
    ws_metas.merge_cells('A1:G1')
    ws_metas['A1'] = 'METAS Y ACTIVIDADES DETALLADAS'
    ws_metas['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_metas['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    
    metas = proyecto.metas.prefetch_related('actividades').all()
    
    for idx_meta, meta in enumerate(metas, 1):
        # Título de la meta
        ws_metas.merge_cells(f'A{row}:G{row}')
        ws_metas.cell(row=row, column=1, value=f'META {idx_meta}: {meta.descripcion}')
        ws_metas.cell(row=row, column=1).font = Font(bold=True, size=12, color="1e40af")
        ws_metas.cell(row=row, column=1).fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        row += 1
        
        # Encabezados de actividades
        headers = ['#', 'Actividad', 'U. Medida', 'Cantidad', 'Recursos', 'Medio Verif.', 'Cuantificable']
        for col, header in enumerate(headers, start=1):
            cell = ws_metas.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row += 1
        
        # Actividades
        for idx_act, actividad in enumerate(meta.actividades.all(), 1):
            ws_metas.cell(row=row, column=1, value=idx_act).border = border
            ws_metas.cell(row=row, column=2, value=actividad.descripcion).border = border
            ws_metas.cell(row=row, column=3, value=actividad.unidad_medida).border = border
            ws_metas.cell(row=row, column=4, value=actividad.cantidad_programada).border = border
            ws_metas.cell(row=row, column=5, value=float(actividad.total_recursos)).border = border
            ws_metas.cell(row=row, column=6, value=actividad.medio_verificacion).border = border
            ws_metas.cell(row=row, column=7, value='Sí' if actividad.es_cuantificable else 'No').border = border
            row += 1
        
        row += 1
    
    ws_metas.column_dimensions['A'].width = 5
    ws_metas.column_dimensions['B'].width = 50
    ws_metas.column_dimensions['C'].width = 15
    ws_metas.column_dimensions['D'].width = 12
    ws_metas.column_dimensions['E'].width = 15
    ws_metas.column_dimensions['F'].width = 30
    ws_metas.column_dimensions['G'].width = 15
    
    # Hoja 4: Programación Mensual
    ws_prog = wb.create_sheet("Programación Mensual")
    
    ws_prog.merge_cells('A1:F1')
    ws_prog['A1'] = 'PROGRAMACIÓN MENSUAL'
    ws_prog['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_prog['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    
    for meta in metas:
        for actividad in meta.actividades.all():
            if actividad.avances.exists():
                # Título de actividad
                ws_prog.merge_cells(f'A{row}:F{row}')
                ws_prog.cell(row=row, column=1, value=f'Actividad: {actividad.descripcion}')
                ws_prog.cell(row=row, column=1).font = Font(bold=True, size=11)
                ws_prog.cell(row=row, column=1).fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
                row += 1
                
                # Encabezados
                prog_headers = ['Mes', 'Programado', 'Realizado', '% Cumplimiento', 'Estado', 'No Planificada']
                for col, header in enumerate(prog_headers, start=1):
                    cell = ws_prog.cell(row=row, column=col)
                    cell.value = header
                    cell.fill = subheader_fill
                    cell.font = subheader_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                row += 1
                
                # Datos mensuales
                for avance in actividad.avances.all().order_by('mes'):
                    estado = 'Excelente' if avance.cumplimiento >= 90 else 'Bueno' if avance.cumplimiento >= 70 else 'Regular' if avance.cumplimiento >= 50 else 'Deficiente'
                    
                    ws_prog.cell(row=row, column=1, value=avance.get_mes_display()).border = border
                    ws_prog.cell(row=row, column=2, value=avance.cantidad_programada_mes).border = border
                    ws_prog.cell(row=row, column=3, value=avance.cantidad_realizada).border = border
                    ws_prog.cell(row=row, column=4, value=f'{avance.cumplimiento}%').border = border
                    ws_prog.cell(row=row, column=5, value=estado).border = border
                    ws_prog.cell(row=row, column=6, value='Sí' if avance.es_no_planificada else 'No').border = border
                    row += 1
                
                row += 1
    
    ws_prog.column_dimensions['A'].width = 15
    ws_prog.column_dimensions['B'].width = 12
    ws_prog.column_dimensions['C'].width = 12
    ws_prog.column_dimensions['D'].width = 18
    ws_prog.column_dimensions['E'].width = 15
    ws_prog.column_dimensions['F'].width = 15
    
    # Hoja 5: Evidencias
    ws_evid = wb.create_sheet("Evidencias")
    
    ws_evid.merge_cells('A1:E1')
    ws_evid['A1'] = 'EVIDENCIAS'
    ws_evid['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_evid['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    
    evid_headers = ['Actividad', 'Tipo', 'Mes', 'Descripción', 'Fecha Subida']
    for col, header in enumerate(evid_headers, start=1):
        cell = ws_evid.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    
    for meta in metas:
        for actividad in meta.actividades.all():
            evidencias = actividad.evidencias.all()
            for evidencia in evidencias:
                ws_evid.cell(row=row, column=1, value=actividad.descripcion[:50]).border = border
                ws_evid.cell(row=row, column=2, value=evidencia.tipo).border = border
                ws_evid.cell(row=row, column=3, value=evidencia.get_mes_display() if evidencia.mes else 'N/A').border = border
                ws_evid.cell(row=row, column=4, value=evidencia.descripcion or 'Sin descripción').border = border
                ws_evid.cell(row=row, column=5, value=evidencia.fecha_subida.strftime('%d/%m/%Y')).border = border
                row += 1
    
    ws_evid.column_dimensions['A'].width = 40
    ws_evid.column_dimensions['B'].width = 15
    ws_evid.column_dimensions['C'].width = 15
    ws_evid.column_dimensions['D'].width = 50
    ws_evid.column_dimensions['E'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="POA_Detallado_{proyecto.unidad.unidad.nombre}_{proyecto.anio}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


def generar_pdf_unidades(unidades_con_rendimiento, usuario):
    """
    Genera un PDF con el reporte de todas las unidades y su cumplimiento
    Usado por administrador y auditor
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Unidades_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elementos = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitulo_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=15,
        fontName='Helvetica-Bold'
    )
    
    # Encabezado
    elementos.append(Paragraph('REPORTE DE UNIDADES Y CUMPLIMIENTO', titulo_style))
    elementos.append(Paragraph(f'Fecha de Generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elementos.append(Spacer(1, 0.3*inch))
    
    # Resumen estadístico
    elementos.append(Paragraph('RESUMEN GENERAL', subtitulo_style))
    
    total_unidades = len(unidades_con_rendimiento)
    excelentes = sum(1 for u in unidades_con_rendimiento if u.rendimiento >= 80)
    buenos = sum(1 for u in unidades_con_rendimiento if 60 <= u.rendimiento < 80)
    regulares = sum(1 for u in unidades_con_rendimiento if 40 <= u.rendimiento < 60)
    bajos = sum(1 for u in unidades_con_rendimiento if u.rendimiento < 40)
    
    resumen_data = [
        ['Categoría', 'Cantidad', 'Porcentaje'],
        ['Total de Unidades', str(total_unidades), '100%'],
        ['Excelente (≥80%)', str(excelentes), f'{(excelentes/total_unidades*100):.1f}%' if total_unidades > 0 else '0%'],
        ['Bueno (60-79%)', str(buenos), f'{(buenos/total_unidades*100):.1f}%' if total_unidades > 0 else '0%'],
        ['Regular (40-59%)', str(regulares), f'{(regulares/total_unidades*100):.1f}%' if total_unidades > 0 else '0%'],
        ['Bajo (<40%)', str(bajos), f'{(bajos/total_unidades*100):.1f}%' if total_unidades > 0 else '0%'],
    ]
    
    resumen_tabla = Table(resumen_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
    resumen_tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c4a6e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#dbeafe')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    elementos.append(resumen_tabla)
    elementos.append(Spacer(1, 0.4*inch))
    
    # Tabla detallada de unidades
    elementos.append(Paragraph('DETALLE POR UNIDAD', subtitulo_style))
    
    unidades_data = [['#', 'Unidad', 'Email', 'Total\nProyectos', 'Aprobados', 'Rendimiento', 'Categoría']]
    
    for idx, unidad in enumerate(unidades_con_rendimiento, 1):
        unidades_data.append([
            str(idx),
            Paragraph(unidad.unidad.nombre, styles['Normal']),
            Paragraph(unidad.email, styles['Normal']),
            str(unidad.total_proyectos),
            str(unidad.count_proyectos_aprobados),
            f'{unidad.rendimiento}%',
            unidad.categoria
        ])
    
    unidades_tabla = Table(unidades_data, colWidths=[0.3*inch, 2*inch, 1.8*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.8*inch])
    
    # Estilos base de la tabla
    tabla_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c4a6e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]
    
    # Agregar colores según categoría
    for idx, unidad in enumerate(unidades_con_rendimiento, 1):
        row = idx
        if unidad.rendimiento >= 80:
            tabla_styles.append(('BACKGROUND', (5, row), (6, row), colors.HexColor('#d1fae5')))
        elif unidad.rendimiento >= 60:
            tabla_styles.append(('BACKGROUND', (5, row), (6, row), colors.HexColor('#fef3c7')))
        elif unidad.rendimiento >= 40:
            tabla_styles.append(('BACKGROUND', (5, row), (6, row), colors.HexColor('#fed7aa')))
        else:
            tabla_styles.append(('BACKGROUND', (5, row), (6, row), colors.HexColor('#fecaca')))
    
    unidades_tabla.setStyle(TableStyle(tabla_styles))
    elementos.append(unidades_tabla)
    
    # Pie de página
    elementos.append(Spacer(1, 0.5*inch))
    elementos.append(Paragraph('_' * 100, styles['Normal']))
    elementos.append(Paragraph(f'Documento generado automáticamente el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}', styles['Normal']))
    elementos.append(Paragraph('Sistema de Gestión POA - Alcaldía', styles['Normal']))
    
    doc.build(elementos)
    return response


def generar_excel_unidades(unidades_con_rendimiento, usuario):
    """
    Genera un Excel con el reporte de todas las unidades y su cumplimiento
    Usado por administrador y auditor
    """
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="0c4a6e", end_color="0c4a6e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    excelente_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    bueno_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    regular_fill = PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid")
    bajo_fill = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Hoja 1: Resumen General
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    
    ws_resumen.merge_cells('A1:D1')
    ws_resumen['A1'] = 'REPORTE DE UNIDADES Y CUMPLIMIENTO'
    ws_resumen['A1'].font = Font(bold=True, size=18, color="0c4a6e")
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    ws_resumen.merge_cells('A2:D2')
    ws_resumen['A2'] = f'Fecha de Generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws_resumen['A2'].alignment = Alignment(horizontal='center')
    
    # Resumen estadístico
    row = 4
    ws_resumen.cell(row=row, column=1, value='RESUMEN ESTADÍSTICO').font = Font(bold=True, size=14, color="0c4a6e")
    ws_resumen.merge_cells(f'A{row}:D{row}')
    
    row += 2
    total_unidades = len(unidades_con_rendimiento)
    excelentes = sum(1 for u in unidades_con_rendimiento if u.rendimiento >= 80)
    buenos = sum(1 for u in unidades_con_rendimiento if 60 <= u.rendimiento < 80)
    regulares = sum(1 for u in unidades_con_rendimiento if 40 <= u.rendimiento < 60)
    bajos = sum(1 for u in unidades_con_rendimiento if u.rendimiento < 40)
    
    resumen_headers = ['Categoría', 'Cantidad', 'Porcentaje', 'Rango']
    for col, header in enumerate(resumen_headers, start=1):
        cell = ws_resumen.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    resumen_data = [
        ('Total de Unidades', total_unidades, '100%', 'Todas'),
        ('Excelente', excelentes, f'{(excelentes/total_unidades*100):.1f}%' if total_unidades > 0 else '0%', '≥80%'),
        ('Bueno', buenos, f'{(buenos/total_unidades*100):.1f}%' if total_unidades > 0 else '0%', '60-79%'),
        ('Regular', regulares, f'{(regulares/total_unidades*100):.1f}%' if total_unidades > 0 else '0%', '40-59%'),
        ('Bajo', bajos, f'{(bajos/total_unidades*100):.1f}%' if total_unidades > 0 else '0%', '<40%'),
    ]
    
    for categoria, cantidad, porcentaje, rango in resumen_data:
        ws_resumen.cell(row=row, column=1, value=categoria).border = border
        ws_resumen.cell(row=row, column=2, value=cantidad).border = border
        ws_resumen.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws_resumen.cell(row=row, column=3, value=porcentaje).border = border
        ws_resumen.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws_resumen.cell(row=row, column=4, value=rango).border = border
        ws_resumen.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        
        if categoria == 'Excelente':
            ws_resumen.cell(row=row, column=1).fill = excelente_fill
        elif categoria == 'Bueno':
            ws_resumen.cell(row=row, column=1).fill = bueno_fill
        elif categoria == 'Regular':
            ws_resumen.cell(row=row, column=1).fill = regular_fill
        elif categoria == 'Bajo':
            ws_resumen.cell(row=row, column=1).fill = bajo_fill
        
        row += 1
    
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    
    # Hoja 2: Detalle por Unidad
    ws_detalle = wb.create_sheet("Detalle por Unidad")
    
    ws_detalle.merge_cells('A1:G1')
    ws_detalle['A1'] = 'DETALLE POR UNIDAD'
    ws_detalle['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_detalle['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    detalle_headers = ['#', 'Unidad', 'Email', 'Total Proyectos', 'Aprobados', 'Rendimiento (%)', 'Categoría']
    for col, header in enumerate(detalle_headers, start=1):
        cell = ws_detalle.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    
    for idx, unidad in enumerate(unidades_con_rendimiento, 1):
        ws_detalle.cell(row=row, column=1, value=idx).border = border
        ws_detalle.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        ws_detalle.cell(row=row, column=2, value=unidad.unidad.nombre).border = border
        ws_detalle.cell(row=row, column=3, value=unidad.email).border = border
        
        ws_detalle.cell(row=row, column=4, value=unidad.total_proyectos).border = border
        ws_detalle.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        
        ws_detalle.cell(row=row, column=5, value=unidad.count_proyectos_aprobados).border = border
        ws_detalle.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        
        ws_detalle.cell(row=row, column=6, value=unidad.rendimiento).border = border
        ws_detalle.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        ws_detalle.cell(row=row, column=6).fill = unidad.fill
        
        ws_detalle.cell(row=row, column=7, value=unidad.categoria).border = border
        ws_detalle.cell(row=row, column=7).alignment = Alignment(horizontal='center')
        ws_detalle.cell(row=row, column=7).fill = unidad.fill
        
        row += 1
    
    ws_detalle.column_dimensions['A'].width = 5
    ws_detalle.column_dimensions['B'].width = 40
    ws_detalle.column_dimensions['C'].width = 35
    ws_detalle.column_dimensions['D'].width = 18
    ws_detalle.column_dimensions['E'].width = 15
    ws_detalle.column_dimensions['F'].width = 18
    ws_detalle.column_dimensions['G'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Unidades_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response








########################################################################################################################
"""
Módulo de utilidades compartidas para exportación de reportes
Usado por administrador y auditor para evitar duplicación de código
"""
from django.http import HttpResponse
from django.db.models import Sum, Avg, Count, Q
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
from decimal import Decimal

# Asumiendo que tus modelos están importados
# from poa.models import Actividad, AvanceMensual, Evidencia, Proyecto, MetaProyecto
# (Los modelos fueron provistos en el prompt)


def generar_pdf_todos_proyectos(proyectos_qs, usuario):
    """
    Genera un PDF consolidado de TODOS los proyectos con toda su información
    Usado por administrador y auditor
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="POA_Consolidado_Todos_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.4*inch, bottomMargin=0.4*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elementos = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados (copiados de tu referencia)
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=15,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitulo_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Heading3'],
        fontSize=11,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    
    small_style = ParagraphStyle(
        'SmallText',
        parent=styles['Normal'],
        fontSize=7,
        leading=9
    )
    
    # Encabezado del documento
    elementos.append(Paragraph('REPORTE CONSOLIDADO DE PROYECTOS POA', titulo_style))
    elementos.append(Paragraph(f'Fecha de Generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elementos.append(Paragraph(f'Generado por: {usuario.email}', styles['Normal']))
    elementos.append(Spacer(1, 0.3*inch))

    # Optimizar la consulta
    proyectos_a_procesar = proyectos_qs.select_related('unidad__unidad', 'aprobado_por').prefetch_related(
        'metas__actividades__avances', 
        'metas__actividades__evidencias'
    )

    for idx, proyecto in enumerate(proyectos_a_procesar):
        if idx > 0:
            elementos.append(PageBreak()) # Empezar cada proyecto en una nueva página

        # Encabezado del proyecto
        elementos.append(Paragraph(f'PROYECTO: {proyecto.nombre.upper()}', titulo_style))
        elementos.append(Paragraph(f'<b>Unidad Responsable:</b> {proyecto.unidad.unidad.nombre}', styles['Normal']))
        elementos.append(Paragraph(f'<b>Año de Ejecución:</b> {proyecto.anio}', styles['Normal']))
        elementos.append(Paragraph(f'<b>Estado:</b> {proyecto.get_estado_display()}', styles['Normal']))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Información general
        elementos.append(Paragraph('INFORMACIÓN GENERAL', subtitulo_style))
        
        info_data = [
            ['Objetivo de la Unidad:', Paragraph(proyecto.objetivo_unidad or 'No especificado', styles['Normal'])],
            ['Fecha de Creación:', proyecto.fecha_creacion.strftime('%d/%m/%Y %H:%M')],
            ['Última Modificación:', proyecto.fecha_modificacion.strftime('%d/%m/%Y %H:%M')],
        ]
        
        if proyecto.estado == 'APROBADO':
            info_data.append(['Aprobado por:', proyecto.aprobado_por.email if proyecto.aprobado_por else 'N/A'])
            info_data.append(['Fecha de Aprobación:', proyecto.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if proyecto.fecha_aprobacion else 'N/A'])
        elif proyecto.estado == 'RECHAZADO':
            info_data.append(['Motivo de Rechazo:', Paragraph(proyecto.motivo_rechazo or 'No especificado', styles['Normal'])])
        
        info_tabla = Table(info_data, colWidths=[2*inch, 4.5*inch])
        info_tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elementos.append(info_tabla)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Resumen de estadísticas (Se ejecuta por proyecto, igual que en la ref)
        total_actividades = Actividad.objects.filter(meta__proyecto=proyecto).count()
        presupuesto_total = Actividad.objects.filter(meta__proyecto=proyecto).aggregate(
            total=Sum('total_recursos')
        )['total'] or Decimal('0.00')
        total_evidencias = Evidencia.objects.filter(actividad__meta__proyecto=proyecto).count()
        
        elementos.append(Paragraph('RESUMEN EJECUTIVO', subtitulo_style))
        
        resumen_data = [
            ['Total de Metas:', str(proyecto.metas.count())],
            ['Total de Actividades:', str(total_actividades)],
            ['Presupuesto Total:', f'${presupuesto_total:,.2f}'],
            ['Total de Evidencias:', str(total_evidencias)],
        ]
        
        resumen_tabla = Table(resumen_data, colWidths=[2*inch, 4.5*inch])
        resumen_tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dbeafe')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elementos.append(resumen_tabla)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Metas y actividades detalladas
        elementos.append(Paragraph('METAS Y ACTIVIDADES DETALLADAS', subtitulo_style))
        elementos.append(Spacer(1, 0.15*inch))
        
        metas = proyecto.metas.all() # El prefetch_related se usará aquí
        
        for idx_meta, meta in enumerate(metas, 1):
            # Título de la meta
            elementos.append(Paragraph(f'META {idx_meta}: {meta.descripcion}', meta_style))
            
            actividades_data = [['#', 'Actividad', 'U.M.', 'Cant.', 'Recursos']]
            
            actividades_meta = meta.actividades.all()
            for idx_act, actividad in enumerate(actividades_meta, 1):
                actividades_data.append([
                    str(idx_act),
                    Paragraph(actividad.descripcion, small_style),
                    Paragraph(actividad.unidad_medida[:12], small_style),
                    str(actividad.cantidad_programada),
                    f'${actividad.total_recursos:,.0f}'
                ])
            
            if len(actividades_data) > 1:
                act_tabla = Table(actividades_data, colWidths=[0.3*inch, 3.5*inch, 0.7*inch, 0.5*inch, 1*inch])
                act_tabla.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c4a6e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ]))
                elementos.append(act_tabla)
                elementos.append(Spacer(1, 0.12*inch))
                
                # Programación mensual para cada actividad
                for actividad in actividades_meta:
                    if actividad.avances.exists():
                        elementos.append(Paragraph(f'Programación: {actividad.descripcion[:60]}...', styles['Heading4']))
                        
                        prog_data = [['Mes', 'Prog.', 'Real.', '%', 'Estado']]
                        
                        for avance in actividad.avances.all().order_by('mes'):
                            estado = 'Excelente' if avance.cumplimiento >= 90 else 'Bueno' if avance.cumplimiento >= 70 else 'Regular' if avance.cumplimiento >= 50 else 'Deficiente'
                            prog_data.append([
                                avance.get_mes_display()[:3],
                                str(avance.cantidad_programada_mes),
                                str(avance.cantidad_realizada),
                                f'{avance.cumplimiento:.0f}%',
                                estado
                            ])
                        
                        prog_tabla = Table(prog_data, colWidths=[0.9*inch, 0.7*inch, 0.7*inch, 0.7*inch, 1*inch])
                        prog_tabla.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 7),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                            ('TOPPADDING', (0, 0), (-1, 0), 5),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eff6ff')),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('FONTSIZE', (0, 1), (-1, -1), 6),
                            ('TOPPADDING', (0, 1), (-1, -1), 3),
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                        ]))
                        elementos.append(prog_tabla)
                        elementos.append(Spacer(1, 0.08*inch))
                        
                        evidencias = actividad.evidencias.all()
                        if evidencias.exists():
                            elementos.append(Paragraph(f'Evidencias ({evidencias.count()})', styles['Heading4']))
                            
                            evid_data = [['Tipo', 'Mes', 'Descripción', 'Fecha']]
                            
                            for evidencia in evidencias[:8]:
                                evid_data.append([
                                    evidencia.tipo[:10],
                                    evidencia.get_mes_display()[:3] if evidencia.mes else 'N/A',
                                    Paragraph(evidencia.descripcion[:60] if evidencia.descripcion else 'Sin descripción', small_style),
                                    evidencia.fecha_subida.strftime('%d/%m/%y')
                                ])
                            
                            evid_tabla = Table(evid_data, colWidths=[0.8*inch, 0.7*inch, 2.8*inch, 0.8*inch])
                            evid_tabla.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0, 0), (-1, 0), 7),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                                ('TOPPADDING', (0, 0), (-1, 0), 5),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf4')),
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                                ('FONTSIZE', (0, 1), (-1, -1), 6),
                                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                                ('TOPPADDING', (0, 1), (-1, -1), 3),
                                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                            ]))
                            elementos.append(evid_tabla)
                            elementos.append(Spacer(1, 0.1*inch))
            
            # Salto de página por META (como en tu función original)
            elementos.append(PageBreak())

    
    # Pie de página al final del documento
    elementos.append(Spacer(1, 0.5*inch))
    elementos.append(Paragraph('_' * 80, styles['Normal']))
    elementos.append(Paragraph(f'Documento generado automáticamente el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}', styles['Normal']))
    elementos.append(Paragraph('Sistema de Gestión POA - Alcaldía', styles['Normal']))
    
    doc.build(elementos)
    return response


def generar_excel_todos_proyectos(proyectos_qs, usuario):
    """
    Genera un Excel consolidado de TODOS los proyectos con toda su información
    Usado por administrador y auditor
    """
    wb = Workbook()
    wb.remove(wb.active) # Eliminar hoja por defecto
    
    # Estilos
    header_fill = PatternFill(start_color="0c4a6e", end_color="0c4a6e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- Hoja 1: Resumen de Proyectos ---
    ws_res = wb.create_sheet("Resumen_Proyectos")
    
    ws_res.merge_cells('A1:K1')
    ws_res['A1'] = 'REPORTE CONSOLIDADO DE PROYECTOS'
    ws_res['A1'].font = Font(bold=True, size=18, color="0c4a6e")
    ws_res['A1'].alignment = Alignment(horizontal='center')
    
    headers_res = ['Unidad', 'Proyecto', 'Año', 'Estado', 'Objetivo Unidad', 'Aprobado por', 'Fecha Aprob.', 'Total Metas', 'Total Actividades', 'Presupuesto Total', 'Total Evidencias']
    
    for col, header in enumerate(headers_res, start=1):
        cell = ws_res.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row_res = 4
    
    # Optimizar con anotaciones para el resumen
    proyectos_con_stats = proyectos_qs.select_related('unidad__unidad', 'aprobado_por').annotate(
        count_metas=Count('metas', distinct=True),
        count_actividades=Count('metas__actividades', distinct=True),
        sum_presupuesto=Sum('metas__actividades__total_recursos'),
        count_evidencias=Count('metas__actividades__evidencias', distinct=True)
    )

    for proyecto in proyectos_con_stats:
        ws_res.cell(row=row_res, column=1, value=proyecto.unidad.unidad.nombre).border = border
        ws_res.cell(row=row_res, column=2, value=proyecto.nombre).border = border
        ws_res.cell(row=row_res, column=3, value=proyecto.anio).border = border
        ws_res.cell(row=row_res, column=4, value=proyecto.get_estado_display()).border = border
        ws_res.cell(row=row_res, column=5, value=proyecto.objetivo_unidad).border = border
        ws_res.cell(row=row_res, column=6, value=proyecto.aprobado_por.email if proyecto.aprobado_por else 'N/A').border = border
        ws_res.cell(row=row_res, column=7, value=proyecto.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if proyecto.fecha_aprobacion else 'N/A').border = border
        ws_res.cell(row=row_res, column=8, value=proyecto.count_metas).border = border
        ws_res.cell(row=row_res, column=9, value=proyecto.count_actividades).border = border
        ws_res.cell(row=row_res, column=10, value=proyecto.sum_presupuesto or 0).border = border
        ws_res.cell(row=row_res, column=10).number_format = '"$"#,##0.00'
        ws_res.cell(row=row_res, column=11, value=proyecto.count_evidencias).border = border
        row_res += 1

    ws_res.column_dimensions['A'].width = 30
    ws_res.column_dimensions['B'].width = 40
    ws_res.column_dimensions['C'].width = 8
    ws_res.column_dimensions['D'].width = 15
    ws_res.column_dimensions['E'].width = 40
    ws_res.column_dimensions['F'].width = 25
    ws_res.column_dimensions['G'].width = 18
    ws_res.column_dimensions['H'].width = 12
    ws_res.column_dimensions['I'].width = 12
    ws_res.column_dimensions['J'].width = 18
    ws_res.column_dimensions['K'].width = 12

    
    # --- Hoja 2: Detalle Metas y Actividades ---
    ws_metas = wb.create_sheet("Detalle_Metas_Actividades")
    
    ws_metas.merge_cells('A1:K1')
    ws_metas['A1'] = 'DETALLE DE METAS Y ACTIVIDADES'
    ws_metas['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_metas['A1'].alignment = Alignment(horizontal='center')
    
    headers_metas = ['Unidad', 'Proyecto', 'Año', 'Meta (Desc)', 'Actividad (Desc)', 'U. Medida', 'Cant. Programada', 'Recursos', 'Total Recursos', 'Medio Verif.', 'Cuantificable']
    
    for col, header in enumerate(headers_metas, start=1):
        cell = ws_metas.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
    row_meta = 4
    
    # Re-usar el queryset optimizado
    proyectos_full = proyectos_qs.select_related('unidad__unidad').prefetch_related('metas__actividades')

    for proyecto in proyectos_full:
        for meta in proyecto.metas.all():
            for actividad in meta.actividades.all():
                ws_metas.cell(row=row_meta, column=1, value=proyecto.unidad.unidad.nombre).border = border
                ws_metas.cell(row=row_meta, column=2, value=proyecto.nombre).border = border
                ws_metas.cell(row=row_meta, column=3, value=proyecto.anio).border = border
                ws_metas.cell(row=row_meta, column=4, value=meta.descripcion).border = border
                ws_metas.cell(row=row_meta, column=5, value=actividad.descripcion).border = border
                ws_metas.cell(row=row_meta, column=6, value=actividad.unidad_medida).border = border
                ws_metas.cell(row=row_meta, column=7, value=actividad.cantidad_programada).border = border
                ws_metas.cell(row=row_meta, column=8, value=actividad.recursos).border = border
                ws_metas.cell(row=row_meta, column=9, value=actividad.total_recursos).border = border
                ws_metas.cell(row=row_meta, column=9).number_format = '"$"#,##0.00'
                ws_metas.cell(row=row_meta, column=10, value=actividad.medio_verificacion).border = border
                ws_metas.cell(row=row_meta, column=11, value='Sí' if actividad.es_cuantificable else 'No').border = border
                row_meta += 1
                
    ws_metas.column_dimensions['A'].width = 30
    ws_metas.column_dimensions['B'].width = 40
    ws_metas.column_dimensions['C'].width = 8
    ws_metas.column_dimensions['D'].width = 40
    ws_metas.column_dimensions['E'].width = 50
    ws_metas.column_dimensions['F'].width = 15
    ws_metas.column_dimensions['G'].width = 12
    ws_metas.column_dimensions['H'].width = 20
    ws_metas.column_dimensions['I'].width = 15
    ws_metas.column_dimensions['J'].width = 30
    ws_metas.column_dimensions['K'].width = 12

    # --- Hoja 3: Detalle Programación ---
    ws_prog = wb.create_sheet("Detalle_Programacion")

    ws_prog.merge_cells('A1:J1')
    ws_prog['A1'] = 'DETALLE DE PROGRAMACIÓN MENSUAL'
    ws_prog['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_prog['A1'].alignment = Alignment(horizontal='center')
    
    headers_prog = ['Unidad', 'Proyecto', 'Actividad', 'Mes', 'Año', 'Prog. Mes', 'Realizado', '% Cumpl.', 'Causal Incumpl.', 'No Planificada']
    
    for col, header in enumerate(headers_prog, start=1):
        cell = ws_prog.cell(row=3, column=col)
        cell.value = header
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
    row_prog = 4
    
    proyectos_prog = proyectos_qs.select_related('unidad__unidad').prefetch_related('metas__actividades__avances')
    
    for proyecto in proyectos_prog:
        for meta in proyecto.metas.all():
            for actividad in meta.actividades.all():
                for avance in actividad.avances.all().order_by('mes'):
                    estado = 'Excelente' if avance.cumplimiento >= 90 else 'Bueno' if avance.cumplimiento >= 70 else 'Regular' if avance.cumplimiento >= 50 else 'Deficiente'
                    
                    ws_prog.cell(row=row_prog, column=1, value=proyecto.unidad.unidad.nombre).border = border
                    ws_prog.cell(row=row_prog, column=2, value=proyecto.nombre).border = border
                    ws_prog.cell(row=row_prog, column=3, value=actividad.descripcion).border = border
                    ws_prog.cell(row=row_prog, column=4, value=avance.get_mes_display()).border = border
                    ws_prog.cell(row=row_prog, column=5, value=avance.anio).border = border
                    ws_prog.cell(row=row_prog, column=6, value=avance.cantidad_programada_mes).border = border
                    ws_prog.cell(row=row_prog, column=7, value=avance.cantidad_realizada).border = border
                    ws_prog.cell(row=row_prog, column=8, value=f'{avance.cumplimiento}%').border = border
                    ws_prog.cell(row=row_prog, column=9, value=avance.causal_incumplimiento).border = border
                    ws_prog.cell(row=row_prog, column=10, value='Sí' if avance.es_no_planificada else 'No').border = border
                    row_prog += 1

    ws_prog.column_dimensions['A'].width = 30
    ws_prog.column_dimensions['B'].width = 40
    ws_prog.column_dimensions['C'].width = 50
    ws_prog.column_dimensions['D'].width = 12
    ws_prog.column_dimensions['E'].width = 8
    ws_prog.column_dimensions['F'].width = 12
    ws_prog.column_dimensions['G'].width = 12
    ws_prog.column_dimensions['H'].width = 12
    ws_prog.column_dimensions['I'].width = 30
    ws_prog.column_dimensions['J'].width = 12
    
    # --- Hoja 4: Detalle Evidencias ---
    ws_evid = wb.create_sheet("Detalle_Evidencias")
    
    ws_evid.merge_cells('A1:I1')
    ws_evid['A1'] = 'DETALLE DE EVIDENCIAS'
    ws_evid['A1'].font = Font(bold=True, size=16, color="0c4a6e")
    ws_evid['A1'].alignment = Alignment(horizontal='center')
    
    headers_evid = ['Unidad', 'Proyecto', 'Actividad', 'Tipo', 'Archivo', 'URL', 'Descripción', 'Mes', 'Fecha Subida']
    
    for col, header in enumerate(headers_evid, start=1):
        cell = ws_evid.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
    row_evid = 4
    
    proyectos_evid = proyectos_qs.select_related('unidad__unidad').prefetch_related('metas__actividades__evidencias')

    for proyecto in proyectos_evid:
        for meta in proyecto.metas.all():
            for actividad in meta.actividades.all():
                for evidencia in actividad.evidencias.all():
                    ws_evid.cell(row=row_evid, column=1, value=proyecto.unidad.unidad.nombre).border = border
                    ws_evid.cell(row=row_evid, column=2, value=proyecto.nombre).border = border
                    ws_evid.cell(row=row_evid, column=3, value=actividad.descripcion).border = border
                    ws_evid.cell(row=row_evid, column=4, value=evidencia.tipo).border = border
                    ws_evid.cell(row=row_evid, column=5, value=str(evidencia.archivo) if evidencia.archivo else 'N/A').border = border
                    ws_evid.cell(row=row_evid, column=6, value=evidencia.url if evidencia.url else 'N/A').border = border
                    ws_evid.cell(row=row_evid, column=7, value=evidencia.descripcion).border = border
                    ws_evid.cell(row=row_evid, column=8, value=evidencia.get_mes_display() if evidencia.mes else 'N/A').border = border
                    ws_evid.cell(row=row_evid, column=9, value=evidencia.fecha_subida.strftime('%d/%m/%Y')).border = border
                    row_evid += 1

    ws_evid.column_dimensions['A'].width = 30
    ws_evid.column_dimensions['B'].width = 40
    ws_evid.column_dimensions['C'].width = 50
    ws_evid.column_dimensions['D'].width = 10
    ws_evid.column_dimensions['E'].width = 30
    ws_evid.column_dimensions['F'].width = 30
    ws_evid.column_dimensions['G'].width = 40
    ws_evid.column_dimensions['H'].width = 12
    ws_evid.column_dimensions['I'].width = 18

    # --- Guardar y devolver respuesta ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="POA_Consolidado_Todos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response







def generar_pdf_reporte_trimestral(unidades_trimestrales, usuario, busqueda):
    """
    Genera un PDF con el reporte de cumplimiento trimestral por unidad
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Trimestral_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elementos = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=15,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitulo_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#0c4a6e'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    # Encabezado
    elementos.append(Paragraph('REPORTE DE CUMPLIMIENTO TRIMESTRAL', titulo_style))
    elementos.append(Paragraph(f'Fecha de Generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elementos.append(Paragraph(f'Generado por: {usuario.email}', styles['Normal']))
    if busqueda:
        elementos.append(Paragraph(f'<b>Filtro de búsqueda:</b> "{busqueda}"', styles['Normal']))
    elementos.append(Spacer(1, 0.3*inch))
    
    # Tabla detallada de unidades
    elementos.append(Paragraph('DETALLE POR UNIDAD', subtitulo_style))
    
    # Definir encabezados de tabla
    unidades_data = [['#', 'Unidad', 'Trimestre 1', 'Trimestre 2', 'Trimestre 3', 'Trimestre 4']]
    
    # Estilos base de la tabla
    tabla_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c4a6e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),           # Alinear #
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),          # Alinear T1-T4
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]

    # Llenar datos y agregar colores
    for idx, unidad in enumerate(unidades_trimestrales, 1):
        fila_datos = [
            str(idx),
            Paragraph(unidad['nombre'], styles['Normal']),
            f"{unidad['t1']}%",
            f"{unidad['t2']}%",
            f"{unidad['t3']}%",
            f"{unidad['t4']}%",
        ]
        unidades_data.append(fila_datos)
        
        # Agregar colores a las celdas T1-T4
        for col_idx, valor in enumerate([unidad['t1'], unidad['t2'], unidad['t3'], unidad['t4']], start=2):
            fill_color = colors.HexColor('#fecaca') # Bajo
            if valor >= 80:
                fill_color = colors.HexColor('#d1fae5') # Excelente
            elif valor >= 60:
                fill_color = colors.HexColor('#fef3c7') # Bueno
            elif valor >= 40:
                fill_color = colors.HexColor('#fed7aa') # Regular
            
            tabla_styles.append(('BACKGROUND', (col_idx, idx), (col_idx, idx), fill_color))

    
    unidades_tabla = Table(unidades_data, colWidths=[0.4*inch, 3.1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    unidades_tabla.setStyle(TableStyle(tabla_styles))
    elementos.append(unidades_tabla)
    
    # Pie de página
    elementos.append(Spacer(1, 0.5*inch))
    elementos.append(Paragraph('_' * 100, styles['Normal']))
    elementos.append(Paragraph(f'Total de unidades reportadas: {len(unidades_trimestrales)}', styles['Normal']))
    
    doc.build(elementos)
    return response


def generar_excel_reporte_trimestral(unidades_trimestrales, usuario, busqueda):
    """
    Genera un Excel con el reporte de cumplimiento trimestral por unidad
    """
    wb = Workbook()
    
    # Estilos (copiados de tu generar_excel_unidades)
    header_fill = PatternFill(start_color="0c4a6e", end_color="0c4a6e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    excelente_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    bueno_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    regular_fill = PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid")
    bajo_fill = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Hoja 1: Detalle por Unidad
    ws_detalle = wb.active
    ws_detalle.title = "Reporte Trimestral"
    
    ws_detalle.merge_cells('A1:F1')
    ws_detalle['A1'] = 'REPORTE DE CUMPLIMIENTO TRIMESTRAL'
    ws_detalle['A1'].font = Font(bold=True, size=18, color="0c4a6e")
    ws_detalle['A1'].alignment = Alignment(horizontal='center')
    
    ws_detalle.merge_cells('A2:F2')
    ws_detalle['A2'] = f'Fecha de Generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws_detalle['A2'].alignment = Alignment(horizontal='center')
    
    if busqueda:
        ws_detalle.merge_cells('A3:F3')
        ws_detalle['A3'] = f'Filtro de búsqueda: "{busqueda}"'
        ws_detalle['A3'].alignment = Alignment(horizontal='center')
    
    row = 5
    detalle_headers = ['#', 'Unidad', 'Trimestre 1 (%)', 'Trimestre 2 (%)', 'Trimestre 3 (%)', 'Trimestre 4 (%)']
    for col, header in enumerate(detalle_headers, start=1):
        cell = ws_detalle.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    
    for idx, unidad in enumerate(unidades_trimestrales, 1):
        ws_detalle.cell(row=row, column=1, value=idx).border = border
        ws_detalle.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        ws_detalle.cell(row=row, column=2, value=unidad['nombre']).border = border
        
        # Añadir valores y aplicar estilos condicionales
        trimestres = [unidad['t1'], unidad['t2'], unidad['t3'], unidad['t4']]
        for col_idx, valor in enumerate(trimestres, start=3):
            cell = ws_detalle.cell(row=row, column=col_idx, value=valor)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.00"%"'
            
            if valor >= 80:
                cell.fill = excelente_fill
            elif valor >= 60:
                cell.fill = bueno_fill
            elif valor >= 40:
                cell.fill = regular_fill
            else:
                cell.fill = bajo_fill
        
        row += 1
    
    ws_detalle.column_dimensions['A'].width = 5
    ws_detalle.column_dimensions['B'].width = 50
    ws_detalle.column_dimensions['C'].width = 18
    ws_detalle.column_dimensions['D'].width = 18
    ws_detalle.column_dimensions['E'].width = 18
    ws_detalle.column_dimensions['F'].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Trimestral_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response