"""
Utilidad para generar archivos Excel de POA (Plan Operativo Anual)
basado en las plantillas existentes del sistema.

Esta versión copia un archivo de plantilla existente y lo modifica con los datos,
garantizando que el formato sea idéntico al original.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil
import os


def generar_poa_excel(unidad, proyectos, objetivo_estrategico):
    """
    Genera un archivo Excel con el formato POA para una unidad específica.
    Usa un archivo de plantilla como base para mantener el formato exacto.
    
    Args:
        unidad: Objeto Unidad
        proyectos: QuerySet de proyectos de la unidad
        objetivo_estrategico: Objeto ObjetivoEstrategico seleccionado
    
    Returns:
        Workbook de openpyxl listo para ser guardado
    """
    
    # Ruta al archivo de plantilla
    import os
    from django.conf import settings
    
    # Usar ruta relativa al módulo actual
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(current_dir, 'plantilla_poa.xlsx')
    
    # Cargar la plantilla
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Actualizar el nombre de la hoja
    ws.title = f"POA {datetime.now().year}"
    
    # === ACTUALIZAR ENCABEZADOS ===
    
    # Agregar labels en columna B
    ws['B5'] = '1'
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C5'] = 'UNIDAD:'
    ws['C5'].font = Font(bold=True, size=12)
    ws['C5'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B6'] = '2'
    ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C6'] = 'OBJETIVO ESTRATÉGICO:'
    ws['C6'].font = Font(bold=True, size=12)
    ws['C6'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B7'] = '3'
    ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C7'] = 'OBJETIVO DE LA UNIDAD:'
    ws['C7'].font = Font(bold=True, size=12)
    ws['C7'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Fila 5: Nombre de la unidad (columna D está fusionada D5:F5)
    if not ws.merged_cells.ranges:
        ws.merge_cells('D5:F5')
    ws['D5'] = unidad.nombre.upper()
    ws['D5'].font = Font(size=11)
    ws['D5'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Fila 6: Objetivo estratégico (columna D está fusionada D6:F6)
    if 'D6:F6' not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells('D6:F6')
    ws['D6'] = objetivo_estrategico.descripcion if objetivo_estrategico else ''
    ws['D6'].font = Font(size=11)
    ws['D6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Fila 7: Objetivos de la Unidad (columna D está fusionada D7:F7)
    if 'D7:F7' not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells('D7:F7')
    
    objetivos_unidad = set()
    for proyecto in proyectos:
        if proyecto.objetivo_unidad:
            objetivos_unidad.add(proyecto.objetivo_unidad)
            
    objetivos_texto = '\n'.join([f"- {obj}" for obj in objetivos_unidad])
    
    ws['D7'] = objetivos_texto
    ws['D7'].font = Font(size=11)
    ws['D7'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    # Ajustar altura de fila según cantidad de objetivos
    ws.row_dimensions[7].height = max(15 * len(objetivos_unidad), 30)
    
    # === LIMPIAR DATOS EXISTENTES ===
    # Eliminar filas de datos de ejemplo (desde fila 14 en adelante)
    # Mantener solo las filas de encabezado (1-13)
    max_row = ws.max_row
    if max_row > 13:
        ws.delete_rows(14, max_row - 13)

    # === GENERAR ENCABEZADOS DE MESES Y TRIMESTRES ===
    # Estilos
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    white_font = Font(bold=True, size=10, color="FFFFFF")
    black_font = Font(bold=True, size=9)
    
    # Configuración de columnas
    # Columna G (7) es donde empieza Enero
    start_col_idx = 7 
    
    meses = [
        'ENE', 'FEB', 'MAR',
        'ABR', 'MAY', 'JUN',
        'JUL', 'AGO', 'SEP',
        'OCT', 'NOV', 'DIC'
    ]
    
    quarters = [
        '(Q1) TRIMESTRE 1',
        '(Q2) TRIMESTRE 2',
        '(Q3) TRIMESTRE 3',
        '(Q4) TRIMESTRE 4'
    ]
    
    current_col = start_col_idx
    
    # Fila 11: Trimestres (cada 3 meses + 1 columna de resumen = 13 columnas por trimestre)
    # Fila 12: Meses (cada mes ocupa 4 columnas) y Headers fijos
    # Fila 13: Sub-headers (Prog, Real, Cump, Verif)
    
    # Headers fijos (B-F)
    fixed_headers = {
        'B12': 'N°',
        'C12': '4.PROYECTOS',
        'D12': '5. METAS',
        'E12': '6. ACTIVIDAD',
        'F12': '7. UNIDAD DE MEDIDA'
    }
    
    for cell_ref, text in fixed_headers.items():
        cell = ws[cell_ref]
        cell.value = text
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = header_border

    # Generar columnas de meses y trimestres
    month_counter = 0
    for q_idx in range(4):
        # Inicio del trimestre
        q_start_col = current_col
        
        # 3 Meses por trimestre
        for m in range(3):
            month_name = meses[month_counter]
            month_counter += 1
            
            # Header del Mes (Fila 12) - Fusiona 4 columnas
            col_letter_start = get_column_letter(current_col)
            col_letter_end = get_column_letter(current_col + 3)
            
            # Desfusionar si es necesario
            try:
                ws.unmerge_cells(f'{col_letter_start}12:{col_letter_end}12')
            except:
                pass
            
            ws.merge_cells(f'{col_letter_start}12:{col_letter_end}12')
            cell_month = ws[f'{col_letter_start}12']
            cell_month.value = month_name
            cell_month.font = white_font
            cell_month.fill = header_fill
            cell_month.alignment = Alignment(horizontal='center', vertical='center')
            cell_month.border = header_border
            
            # Sub-headers (Fila 13)
            sub_headers = ['Programado', 'Realizado', 'Cumplimiento', 'Medio de verificación\no\nCausal de\nIncumplimiento']
            for i, sub in enumerate(sub_headers):
                cell_sub = ws.cell(row=13, column=current_col + i)
                cell_sub.value = sub
                cell_sub.font = black_font
                cell_sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_sub.border = header_border
                
                # Ajustar ancho de columna
                col_letter = get_column_letter(current_col + i)
                if i == 3: # Verificación
                    ws.column_dimensions[col_letter].width = 25
                else:
                    ws.column_dimensions[col_letter].width = 12
            
            current_col += 4
            
        # Columna de Seguimiento Trimestral
        col_letter_q = get_column_letter(current_col)
        
        # Header Trimestre (Fila 11) - Cubre todo el trimestre anterior
        q_end_col_letter = get_column_letter(current_col) # Incluye la columna de resumen
        q_start_col_letter = get_column_letter(q_start_col)
        
        try:
            ws.unmerge_cells(f'{q_start_col_letter}11:{q_end_col_letter}11')
        except:
            pass
        
        ws.merge_cells(f'{q_start_col_letter}11:{q_end_col_letter}11')
        cell_q = ws[f'{q_start_col_letter}11']
        cell_q.value = quarters[q_idx]
        cell_q.font = Font(bold=True, size=11)
        cell_q.alignment = Alignment(horizontal='center', vertical='center')
        cell_q.border = header_border
        
        # Header Columna Resumen (Fila 12)
        cell_resumen = ws[f'{col_letter_q}12']
        cell_resumen.value = "Seguimiento\nTrimestral"
        cell_resumen.font = white_font
        cell_resumen.fill = header_fill
        cell_resumen.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_resumen.border = header_border
        
        # Fila 13 vacía o con estilo para resumen
        cell_resumen_sub = ws[f'{col_letter_q}13']
        cell_resumen_sub.border = header_border
        
        ws.column_dimensions[col_letter_q].width = 15
        
        current_col += 1

    # === AGREGAR ENCABEZADOS FINALES (SEMESTRALES, ANUAL, RECURSOS) ===
    
    # Lista de encabezados extra
    extra_headers = [
        ("Seguimiento\nSemestral 1", 15),
        ("Seguimiento\nSemestral 2", 15),
        ("Promedio\nAnual", 15),
        ("Total de\nRecursos $", 20)
    ]
    
    for header_text, width in extra_headers:
        col_letter = get_column_letter(current_col)
        
        # Fila 12
        cell = ws[f'{col_letter}12']
        cell.value = header_text
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = header_border
        
        # Fila 13 (estilo vacío)
        cell_sub = ws[f'{col_letter}13']
        cell_sub.border = header_border
        
        ws.column_dimensions[col_letter].width = width
        current_col += 1
    
    # === AGREGAR DATOS DE PROYECTOS ===
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    fila_actual = 14
    numero_actividad = 1
    
    for proyecto in proyectos:
        fila_inicio_proyecto = fila_actual
        
        for meta in proyecto.metas.all():
            fila_inicio_meta = fila_actual
            
            for actividad in meta.actividades.all():
                # Columna B: Número de actividad
                ws[f'B{fila_actual}'] = numero_actividad
                ws[f'B{fila_actual}'].border = thin_border
                ws[f'B{fila_actual}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Columna E: Actividad
                ws[f'E{fila_actual}'] = actividad.descripcion
                ws[f'E{fila_actual}'].border = thin_border
                ws[f'E{fila_actual}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                
                # Columna F: Unidad de medida
                ws[f'F{fila_actual}'] = actividad.unidad_medida
                ws[f'F{fila_actual}'].border = thin_border
                ws[f'F{fila_actual}'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                # Obtener avances mensuales
                avances = {a.mes: a for a in actividad.avances.all()}
                
                # Llenar datos mensuales
                col_idx = 7 # G
                
                # Guardar referencias para promedios
                cump_cells_q1 = []
                cump_cells_q2 = []
                cump_cells_q3 = []
                cump_cells_q4 = []
                
                for mes in range(1, 13):
                    avance = avances.get(mes)
                    
                    # Columnas: Prog, Real, Cump, Verif
                    col_prog = get_column_letter(col_idx)
                    col_real = get_column_letter(col_idx + 1)
                    col_cump = get_column_letter(col_idx + 2)
                    col_verif = get_column_letter(col_idx + 3)
                    
                    cell_prog = ws[f'{col_prog}{fila_actual}']
                    cell_real = ws[f'{col_real}{fila_actual}']
                    cell_cump = ws[f'{col_cump}{fila_actual}']
                    cell_verif = ws[f'{col_verif}{fila_actual}']
                    
                    # Estilos
                    for cell in [cell_prog, cell_real, cell_cump, cell_verif]:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.font = Font(size=9)
                    
                    # Valores
                    prog_val = 0
                    real_val = 0
                    
                    if avance:
                        prog_val = avance.cantidad_programada_mes
                        real_val = avance.cantidad_realizada
                        
                        if prog_val > 0:
                            cell_prog.value = prog_val
                            cell_real.value = real_val
                            
                            # Fórmula de cumplimiento
                            # =IFERROR(IF(Real/Prog<=1,Real/Prog,"..."),"Actividad no programada")
                            formula = f'=IFERROR(IF({col_real}{fila_actual}/{col_prog}{fila_actual}<=1,{col_real}{fila_actual}/{col_prog}{fila_actual},"El valor excede al 100%, colocar el valor programado y trasladar el excedente a actividades no programadas en este mes"),"Actividad no programada")'
                            cell_cump.value = formula
                            cell_cump.number_format = '0%'
                            
                            # Agregar a lista para promedios
                            ref = f'{col_cump}{fila_actual}'
                            if 1 <= mes <= 3: cump_cells_q1.append(ref)
                            elif 4 <= mes <= 6: cump_cells_q2.append(ref)
                            elif 7 <= mes <= 9: cump_cells_q3.append(ref)
                            elif 10 <= mes <= 12: cump_cells_q4.append(ref)
                            
                        else:
                            cell_prog.value = ""
                            cell_real.value = ""
                            cell_cump.value = "Actividad no programada"
                            cell_cump.font = Font(size=8, italic=True, color="808080")
                    
                    # Avanzar 4 columnas
                    col_idx += 4
                    
                    # Si es fin de trimestre, agregar columna de resumen
                    if mes % 3 == 0:
                        col_resumen = get_column_letter(col_idx)
                        cell_resumen = ws[f'{col_resumen}{fila_actual}']
                        cell_resumen.border = thin_border
                        cell_resumen.alignment = Alignment(horizontal='center', vertical='center')
                        cell_resumen.font = Font(bold=True, size=9)
                        cell_resumen.number_format = '0%'
                        
                        # Fórmula promedio trimestral
                        cells_to_avg = []
                        if mes == 3: cells_to_avg = cump_cells_q1
                        elif mes == 6: cells_to_avg = cump_cells_q2
                        elif mes == 9: cells_to_avg = cump_cells_q3
                        elif mes == 12: cells_to_avg = cump_cells_q4
                        
                        if cells_to_avg:
                            # =IFERROR(AVERAGE(Cump1, Cump2, Cump3),"VALORES NO COLOCADOS")
                            avg_args = ",".join(cells_to_avg)
                            formula_avg = f'=IFERROR(AVERAGE({avg_args}),"VALORES NO COLOCADOS")'
                            cell_resumen.value = formula_avg
                        else:
                            cell_resumen.value = "-"
                            
                        col_idx += 1
                
                # === CÁLCULOS FINALES (S1, S2, ANUAL, RECURSOS) ===
                
                # S1 (Semestre 1)
                col_s1 = get_column_letter(col_idx)
                cell_s1 = ws[f'{col_s1}{fila_actual}']
                cell_s1.border = thin_border
                cell_s1.alignment = Alignment(horizontal='center', vertical='center')
                cell_s1.number_format = '0%'
                
                cells_s1 = cump_cells_q1 + cump_cells_q2
                if cells_s1:
                    avg_args = ",".join(cells_s1)
                    cell_s1.value = f'=IFERROR(AVERAGE({avg_args}),"VALORES NO COLOCADOS")'
                else:
                    cell_s1.value = "-"
                col_idx += 1
                
                # S2 (Semestre 2)
                col_s2 = get_column_letter(col_idx)
                cell_s2 = ws[f'{col_s2}{fila_actual}']
                cell_s2.border = thin_border
                cell_s2.alignment = Alignment(horizontal='center', vertical='center')
                cell_s2.number_format = '0%'
                
                cells_s2 = cump_cells_q3 + cump_cells_q4
                if cells_s2:
                    avg_args = ",".join(cells_s2)
                    cell_s2.value = f'=IFERROR(AVERAGE({avg_args}),"VALORES NO COLOCADOS")'
                else:
                    cell_s2.value = "-"
                col_idx += 1
                
                # Promedio Anual
                col_anual = get_column_letter(col_idx)
                cell_anual = ws[f'{col_anual}{fila_actual}']
                cell_anual.border = thin_border
                cell_anual.alignment = Alignment(horizontal='center', vertical='center')
                cell_anual.number_format = '0%'
                cell_anual.font = Font(bold=True)
                
                cells_total = cells_s1 + cells_s2
                if cells_total:
                    avg_args = ",".join(cells_total)
                    cell_anual.value = f'=IFERROR(AVERAGE({avg_args}),"VALORES NO COLOCADOS")'
                else:
                    cell_anual.value = "-"
                col_idx += 1
                
                # Total Recursos
                col_recursos = get_column_letter(col_idx)
                cell_recursos = ws[f'{col_recursos}{fila_actual}']
                cell_recursos.border = thin_border
                cell_recursos.alignment = Alignment(horizontal='right', vertical='center')
                cell_recursos.number_format = '"$" #,##0.00'
                
                cell_recursos.value = actividad.total_recursos
                col_idx += 1
                
                numero_actividad += 1
                fila_actual += 1
            
            # Fusionar celdas de meta
            if fila_actual > fila_inicio_meta:
                if fila_actual - fila_inicio_meta > 1:
                    ws.merge_cells(f'D{fila_inicio_meta}:D{fila_actual - 1}')
                ws[f'D{fila_inicio_meta}'] = meta.descripcion
                ws[f'D{fila_inicio_meta}'].border = thin_border
                ws[f'D{fila_inicio_meta}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Fusionar celdas de proyecto
        if fila_actual > fila_inicio_proyecto:
            if fila_actual - fila_inicio_proyecto > 1:
                ws.merge_cells(f'C{fila_inicio_proyecto}:C{fila_actual - 1}')
            ws[f'C{fila_inicio_proyecto}'] = proyecto.nombre or f"Proyecto {proyecto.anio} (Sin nombre)"
            ws[f'C{fila_inicio_proyecto}'].border = thin_border
            ws[f'C{fila_inicio_proyecto}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            
    # === AGREGAR ACTIVIDADES NO PLANIFICADAS ===
    # Buscar proyecto no planificado de esta unidad y año
    # unidad es el objeto Unidad (modelo login.Unidad), pero necesitamos el Usuario (login.Usuario)
    # que tiene rol='UNIDAD' y apunta a esta unidad.
    # SIN EMBARGO, en la vista pasamos 'unidad' que es el objeto Usuario (rol=UNIDAD).
    # Revisemos la llamada en views.py: generar_poa_excel(unidad_usuario.unidad, ...)
    # Ah, pasamos unidad_usuario.unidad (objeto Unidad).
    # Pero el proyecto se vincula con Usuario (unidad=Usuario).
    
    # Necesitamos encontrar el Usuario asociado a esta Unidad.
    from login.models import Usuario
    usuario_unidad = Usuario.objects.filter(unidad=unidad, rol='UNIDAD').first()
    
    if usuario_unidad:
        proyecto_no_planificado = proyectos.model.objects.filter(
            unidad=usuario_unidad,
            anio=proyectos.first().anio if proyectos.exists() else datetime.now().year,
            es_no_planificado=True
        ).first()
        
        if proyecto_no_planificado:
            # Título de sección
            fila_actual += 1
            ws.merge_cells(f'B{fila_actual}:F{fila_actual}')
            ws[f'B{fila_actual}'] = "ACTIVIDADES NO PLANIFICADAS"
            ws[f'B{fila_actual}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'B{fila_actual}'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            ws[f'B{fila_actual}'].alignment = Alignment(horizontal='center', vertical='center')
            fila_actual += 1
            
            fila_inicio_proyecto = fila_actual
            
            for meta in proyecto_no_planificado.metas.all():
                fila_inicio_meta = fila_actual
                
                for actividad in meta.actividades.all():
                    # Columna B: Número de actividad
                    ws[f'B{fila_actual}'] = numero_actividad
                    ws[f'B{fila_actual}'].border = thin_border
                    ws[f'B{fila_actual}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Columna E: Actividad
                    ws[f'E{fila_actual}'] = actividad.descripcion
                    ws[f'E{fila_actual}'].border = thin_border
                    ws[f'E{fila_actual}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    
                    # Columna F: Unidad de medida
                    ws[f'F{fila_actual}'] = actividad.unidad_medida
                    ws[f'F{fila_actual}'].border = thin_border
                    ws[f'F{fila_actual}'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    
                    # Obtener avances mensuales
                    avances = {a.mes: a for a in actividad.avances.all()}
                    
                    # Llenar datos mensuales
                    col_idx = 7 # G
                    
                    # Guardar referencias para promedios (aunque en no planificadas suele ser 100% o 0)
                    cump_cells_q1 = []
                    cump_cells_q2 = []
                    cump_cells_q3 = []
                    cump_cells_q4 = []
                    
                    for mes in range(1, 13):
                        avance = avances.get(mes)
                        
                        col_prog = get_column_letter(col_idx)
                        col_real = get_column_letter(col_idx + 1)
                        col_cump = get_column_letter(col_idx + 2)
                        col_verif = get_column_letter(col_idx + 3)
                        
                        cell_prog = ws[f'{col_prog}{fila_actual}']
                        cell_real = ws[f'{col_real}{fila_actual}']
                        cell_cump = ws[f'{col_cump}{fila_actual}']
                        cell_verif = ws[f'{col_verif}{fila_actual}']
                        
                        for cell in [cell_prog, cell_real, cell_cump, cell_verif]:
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.font = Font(size=9)
                        
                        if avance and avance.cantidad_realizada > 0:
                            cell_prog.value = 0 # No planificada no tiene programado
                            cell_real.value = avance.cantidad_realizada
                            cell_cump.value = 1 # 100% si se hizo algo
                            cell_cump.number_format = '0%'
                            
                            ref = f'{col_cump}{fila_actual}'
                            if 1 <= mes <= 3: cump_cells_q1.append(ref)
                            elif 4 <= mes <= 6: cump_cells_q2.append(ref)
                            elif 7 <= mes <= 9: cump_cells_q3.append(ref)
                            elif 10 <= mes <= 12: cump_cells_q4.append(ref)
                        else:
                            cell_prog.value = ""
                            cell_real.value = ""
                            cell_cump.value = "-"
                        
                        col_idx += 4
                        
                        # Resumen trimestral
                        if mes % 3 == 0:
                            col_resumen = get_column_letter(col_idx)
                            cell_resumen = ws[f'{col_resumen}{fila_actual}']
                            cell_resumen.border = thin_border
                            cell_resumen.alignment = Alignment(horizontal='center', vertical='center')
                            cell_resumen.font = Font(bold=True, size=9)
                            cell_resumen.number_format = '0%'
                            
                            cells_to_avg = []
                            if mes == 3: cells_to_avg = cump_cells_q1
                            elif mes == 6: cells_to_avg = cump_cells_q2
                            elif mes == 9: cells_to_avg = cump_cells_q3
                            elif mes == 12: cells_to_avg = cump_cells_q4
                            
                            if cells_to_avg:
                                avg_args = ",".join(cells_to_avg)
                                cell_resumen.value = f'=IFERROR(AVERAGE({avg_args}),"VALORES NO COLOCADOS")'
                            else:
                                cell_resumen.value = "-"
                            col_idx += 1
                    
                    # Cálculos finales
                    # S1
                    col_s1 = get_column_letter(col_idx)
                    cell_s1 = ws[f'{col_s1}{fila_actual}']
                    cell_s1.border = thin_border
                    cell_s1.alignment = Alignment(horizontal='center', vertical='center')
                    cell_s1.number_format = '0%'
                    cells_s1 = cump_cells_q1 + cump_cells_q2
                    if cells_s1:
                        cell_s1.value = f'=IFERROR(AVERAGE({",".join(cells_s1)}),"VALORES NO COLOCADOS")'
                    else:
                        cell_s1.value = "-"
                    col_idx += 1
                    
                    # S2
                    col_s2 = get_column_letter(col_idx)
                    cell_s2 = ws[f'{col_s2}{fila_actual}']
                    cell_s2.border = thin_border
                    cell_s2.alignment = Alignment(horizontal='center', vertical='center')
                    cell_s2.number_format = '0%'
                    cells_s2 = cump_cells_q3 + cump_cells_q4
                    if cells_s2:
                        cell_s2.value = f'=IFERROR(AVERAGE({",".join(cells_s2)}),"VALORES NO COLOCADOS")'
                    else:
                        cell_s2.value = "-"
                    col_idx += 1
                    
                    # Anual
                    col_anual = get_column_letter(col_idx)
                    cell_anual = ws[f'{col_anual}{fila_actual}']
                    cell_anual.border = thin_border
                    cell_anual.alignment = Alignment(horizontal='center', vertical='center')
                    cell_anual.number_format = '0%'
                    cell_anual.font = Font(bold=True)
                    cells_total = cells_s1 + cells_s2
                    if cells_total:
                        cell_anual.value = f'=IFERROR(AVERAGE({",".join(cells_total)}),"VALORES NO COLOCADOS")'
                    else:
                        cell_anual.value = "-"
                    col_idx += 1
                    
                    # Recursos
                    col_recursos = get_column_letter(col_idx)
                    cell_recursos = ws[f'{col_recursos}{fila_actual}']
                    cell_recursos.border = thin_border
                    cell_recursos.alignment = Alignment(horizontal='right', vertical='center')
                    cell_recursos.number_format = '"$" #,##0.00'
                    cell_recursos.value = actividad.total_recursos
                    col_idx += 1
                    
                    numero_actividad += 1
                    fila_actual += 1
                
                # Fusionar celdas de meta
                if fila_actual > fila_inicio_meta:
                    if fila_actual - fila_inicio_meta > 1:
                        ws.merge_cells(f'D{fila_inicio_meta}:D{fila_actual - 1}')
                    ws[f'D{fila_inicio_meta}'] = meta.descripcion
                    ws[f'D{fila_inicio_meta}'].border = thin_border
                    ws[f'D{fila_inicio_meta}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            
            # Fusionar celdas de proyecto (ACTIVIDADES NO PLANIFICADAS)
            if fila_actual > fila_inicio_proyecto:
                if fila_actual - fila_inicio_proyecto > 1:
                    ws.merge_cells(f'C{fila_inicio_proyecto}:C{fila_actual - 1}')
                ws[f'C{fila_inicio_proyecto}'] = "ACTIVIDADES NO PLANIFICADAS"
                ws[f'C{fila_inicio_proyecto}'].border = thin_border
                ws[f'C{fila_inicio_proyecto}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                ws[f'C{fila_inicio_proyecto}'].font = Font(bold=True)
    
    # === CÁLCULO DE PORCENTAJES TRIMESTRALES GLOBALES (80/20) ===
    # Iterar sobre las columnas de resumen trimestral para calcular el promedio global
    # Se debe hacer al final porque necesitamos los datos de todas las filas
    
    # Identificar si existe proyecto no planificado
    proyecto_no_planificado_existe = False
    if usuario_unidad:
        proyecto_no_planificado_existe = proyectos.model.objects.filter(
            unidad=usuario_unidad,
            anio=proyectos.first().anio if proyectos.exists() else datetime.now().year,
            es_no_planificado=True
        ).exists()

    # Columnas de inicio de cada trimestre (basado en la lógica de generación)
    # Q1: Start=7 (G), End=18 (R), Resumen=19 (S) -> No, wait.
    # Logic:
    # Q1: 3 months * 4 cols = 12 cols. Start 7. End 18. Resumen 19.
    # Q2: Start 20. End 31. Resumen 32.
    # Q3: Start 33. End 44. Resumen 45.
    # Q4: Start 46. End 57. Resumen 58.
    
    # Recalculamos las columnas de resumen para estar seguros
    col_resumen_q1 = 7 + (3 * 4) # 19 (S)
    col_resumen_q2 = col_resumen_q1 + 1 + (3 * 4) # 32 (AF)
    col_resumen_q3 = col_resumen_q2 + 1 + (3 * 4) # 45 (AS)
    col_resumen_q4 = col_resumen_q3 + 1 + (3 * 4) # 58 (BF)
    
    cols_resumen = [col_resumen_q1, col_resumen_q2, col_resumen_q3, col_resumen_q4]
    
    # Filas de datos: desde 14 hasta fila_actual - 1
    # Necesitamos distinguir filas de actividades normales vs no planificadas
    # Las no planificadas están al final, después del título "ACTIVIDADES NO PLANIFICADAS"
    
    # Encontrar fila de inicio de no planificadas (si existe)
    fila_inicio_no_planificadas = -1
    for r in range(14, fila_actual):
        cell_b = ws[f'B{r}']
        if cell_b.value == "ACTIVIDADES NO PLANIFICADAS":
            fila_inicio_no_planificadas = r
            break
            
    for i, col_idx in enumerate(cols_resumen):
        col_letter = get_column_letter(col_idx)
        
        # Recolectar valores de cumplimiento
        normal_values = []
        unplanned_values = []
        unplanned_realized = False
        
        for r in range(14, fila_actual):
            # Saltar filas de encabezados/títulos intermedios
            if ws[f'B{r}'].value == "ACTIVIDADES NO PLANIFICADAS":
                continue
                
            # Determinar si es fila de actividad (tiene número en B)
            if not isinstance(ws[f'B{r}'].value, int):
                continue
                
            # Obtener valor de la celda de resumen trimestral de esta fila
            # OJO: La celda de resumen tiene una FÓRMULA.
            # No podemos leer el valor calculado sin evaluar la fórmula o usar una librería que lo haga.
            # PERO, podemos reconstruir el valor buscando las celdas de cumplimiento de los meses.
            
            # Mejor estrategia: Mirar las celdas de cumplimiento de los 3 meses anteriores a esta columna de resumen
            # Columna Resumen es col_idx.
            # Mes 3: col_idx - 1 (Verif), -2 (Cump) -> col_idx - 2
            # Mes 2: col_idx - 1 - 4 (Verif) - 2 (Cump) -> col_idx - 6
            # Mes 1: col_idx - 1 - 4 - 4 - 2 -> col_idx - 10
            
            cump_cols = [col_idx - 2, col_idx - 6, col_idx - 10]
            
            row_values = []
            for c_idx in cump_cols:
                c_letter = get_column_letter(c_idx)
                cell_val = ws[f'{c_letter}{r}'].value
                
                # cell_val puede ser fórmula, número, o string "Actividad no programada"
                # Si es fórmula, asumimos que es válida si Programado > 0.
                # Si es "Actividad no programada" o "-", ignorar.
                
                # Para simplificar, miramos las columnas de Programado y Realizado
                # Prog: c_idx - 2. Real: c_idx - 1.
                prog_cell = ws[f'{get_column_letter(c_idx-2)}{r}'].value
                real_cell = ws[f'{get_column_letter(c_idx-1)}{r}'].value
                
                if isinstance(prog_cell, (int, float)) and prog_cell > 0:
                    if isinstance(real_cell, (int, float)):
                        # Calcular cumplimiento
                        val = min(real_cell / prog_cell, 1.0)
                        row_values.append(val)
                elif fila_inicio_no_planificadas != -1 and r > fila_inicio_no_planificadas:
                     # Es no planificada
                     if isinstance(real_cell, (int, float)) and real_cell > 0:
                         row_values.append(1.0) # 100% si se hizo
                         unplanned_realized = True
            
            if row_values:
                avg_row = sum(row_values) / len(row_values)
                
                if fila_inicio_no_planificadas != -1 and r > fila_inicio_no_planificadas:
                    unplanned_values.append(avg_row)
                else:
                    normal_values.append(avg_row)

        # Calcular promedios globales
        avg_normal = sum(normal_values) / len(normal_values) if normal_values else 0.0
        # avg_unplanned = sum(unplanned_values) / len(unplanned_values) if unplanned_values else 0.0
        
        # Lógica 80/20
        final_score = 0.0
        
        if not proyecto_no_planificado_existe:
            # Caso 1: No existe proyecto no planificado -> 100% Normal
            final_score = avg_normal
        else:
            # Caso 2: Existe proyecto no planificado
            if unplanned_realized:
                # Se hizo al menos una no planificada -> 80% Normal + 20% (Full)
                # User: "volvera a tomar el 100%" -> Implica que el componente no planificado vale 20% completo si se cumple la condición
                final_score = (avg_normal * 0.8) + 0.2
            else:
                # No se hizo ninguna no planificada -> 80% Normal
                final_score = avg_normal * 0.8
        
        # Escribir en el encabezado (Fila 11)
        # Necesitamos ajustar el merge de la fila 11 para hacer espacio
        # Actualmente Fila 11 está mergeada desde Start hasta End (antes de Resumen)
        # La columna Resumen (col_letter) está libre en Fila 11?
        # Revisemos línea 216: ws.merge_cells(f'{q_start_col_letter}11:{q_end_col_letter}11')
        # q_end_col_letter INCLUYE la columna de resumen.
        # Entonces debemos des-mergear y re-mergear.
        
        q_end_col_idx = col_idx
        q_start_col_idx = col_idx - 12
        
        q_start_letter = get_column_letter(q_start_col_idx)
        q_end_letter = get_column_letter(q_end_col_idx)
        
        try:
            ws.unmerge_cells(f'{q_start_letter}11:{q_end_letter}11')
        except:
            pass
            
        # Merge título (hasta penúltima columna)
        q_title_end_letter = get_column_letter(q_end_col_idx - 1)
        ws.merge_cells(f'{q_start_letter}11:{q_title_end_letter}11')
        
        # Restaurar estilo título
        cell_title = ws[f'{q_start_letter}11']
        cell_title.value = quarters[i]
        cell_title.font = Font(bold=True, size=11)
        cell_title.alignment = Alignment(horizontal='center', vertical='center')
        cell_title.border = header_border
        
        # Celda de Porcentaje Global (Última columna del trimestre, Fila 11)
        cell_score = ws[f'{col_letter}11']
        cell_score.value = final_score
        cell_score.number_format = '0%'
        cell_score.font = Font(bold=True, size=11)
        cell_score.alignment = Alignment(horizontal='center', vertical='center')
        cell_score.border = header_border
        
        # Colorear según valor (opcional, pero ayuda)
        if final_score >= 0.8:
            cell_score.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
        elif final_score >= 0.5:
            cell_score.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo
        else:
            cell_score.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rojo

    return wb
