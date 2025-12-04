from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from login.models import Usuario, Unidad
from poa.models import Proyecto, MetaProyecto, Actividad, AvanceMensual, Evidencia, AuditoriaLog
from poa.forms import FormularioProyecto, FormularioMeta, FormularioActividad
from .decorators import admin_required
from openpyxl.cell.cell import Cell
import json
from django.core.paginator import Paginator
from decimal import Decimal
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime

from utils.exportacion import (
    generar_pdf_proyecto_detalle,
    generar_excel_proyecto_detalle,
    generar_pdf_unidades,
    generar_excel_unidades,
    generar_pdf_reporte_trimestral, 
    generar_excel_reporte_trimestral
)


def verificar_admin(user):
    """Verifica que el usuario sea administrador"""
    return user.is_authenticated and user.rol == 'ADMIN'

@login_required
def dashboard_admin(request):
    """Dashboard principal del administrador"""
    if request.user.rol != 'ADMIN':
        return redirect('login:login')
    
    # Estadísticas generales
    total_unidades = Usuario.objects.filter(rol='UNIDAD').count()
    total_proyectos = Proyecto.objects.count()
    proyectos_aprobados = Proyecto.objects.filter(estado='APROBADO').count()
    proyectos_enviados = Proyecto.objects.filter(estado='ENVIADO').count()
    proyectos_borrador = Proyecto.objects.filter(estado='BORRADOR').count()
    
    # Proyectos pendientes de revisión (Enviados)
    pendientes = Proyecto.objects.filter(estado='ENVIADO').select_related('unidad').order_by('-fecha_modificacion')
    
    context = {
        'total_unidades': total_unidades,
        'total_proyectos': total_proyectos,
        'proyectos_aprobados': proyectos_aprobados,
        'proyectos_pendientes': proyectos_enviados,
        'proyectos_borrador': proyectos_borrador,
        'proyectos_revision': pendientes,
    }
    return render(request, 'administrador/dashboard.html', context)



def _get_datos_trimestrales(busqueda_str=""):
    """
    Función de ayuda para calcular y filtrar los datos trimestrales.
    Usada por estadisticas_admin y las vistas de exportación.
    """
    unidades_con_datos = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    unidades_trimestrales_data = []

    for unidad in unidades_con_datos:
        nombre_unidad = unidad.unidad.nombre
        
        if busqueda_str and busqueda_str.lower() not in nombre_unidad.lower():
            continue 
        
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        
        t1_avg, t2_avg, t3_avg, t4_avg = 0.0, 0.0, 0.0, 0.0
        
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            
            t1_data = avances.filter(mes__in=[1, 2, 3]).aggregate(avg=Avg('cumplimiento'))
            t1_avg = float(t1_data['avg'] or 0.0)
            
            t2_data = avances.filter(mes__in=[4, 5, 6]).aggregate(avg=Avg('cumplimiento'))
            t2_avg = float(t2_data['avg'] or 0.0)
            
            t3_data = avances.filter(mes__in=[7, 8, 9]).aggregate(avg=Avg('cumplimiento'))
            t3_avg = float(t3_data['avg'] or 0.0)
            
            t4_data = avances.filter(mes__in=[10, 11, 12]).aggregate(avg=Avg('cumplimiento'))
            t4_avg = float(t4_data['avg'] or 0.0)
            
        unidades_trimestrales_data.append({
            'nombre': nombre_unidad,
            't1': round(t1_avg, 2),
            't2': round(t2_avg, 2),
            't3': round(t3_avg, 2),
            't4': round(t4_avg, 2),
        })
        
    return unidades_trimestrales_data


@admin_required
def estadisticas_admin(request):
    """Página de estadísticas y gráficos del administrador"""
    
    busqueda = request.GET.get('buscar', '')
    page_number = request.GET.get('page', 1)
    active_tab = request.GET.get('tab', 'general')
    
    unidades_con_datos = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    unidades_data = { 'nombres': [], 'cumplimiento': [], 'proyectos': [], 'metricas': [] }
    
    for unidad in unidades_con_datos:
        unidades_data['nombres'].append(unidad.unidad.nombre)
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        cumplimiento_promedio = 0.0
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            resultado_avg = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg']
            if resultado_avg is not None: cumplimiento_promedio = float(resultado_avg)
            unidades_data['cumplimiento'].append(round(cumplimiento_promedio, 2))
            total_evidencias = Evidencia.objects.filter(actividad__meta__proyecto=poa_aprobado).count()
            total_actividades = Actividad.objects.filter(meta__proyecto=poa_aprobado).count()
            unidades_data['metricas'].append([round(cumplimiento_promedio, 2), min(total_evidencias * 10, 100), min(total_actividades * 5, 100), 85, 90])
        else:
            unidades_data['cumplimiento'].append(0.0)
            unidades_data['metricas'].append([0, 0, 0, 0, 0])
        total_proyectos_unidad = Proyecto.objects.filter(unidad=unidad).count()
        unidades_data['proyectos'].append(total_proyectos_unidad)
    
   
    unidades_trimestrales_filtradas = _get_datos_trimestrales(busqueda)
        
    # 2. Paginamos la lista filtrada
    paginator = Paginator(unidades_trimestrales_filtradas, 6) # 6 unidades por página
    page_obj = paginator.get_page(page_number)

    # =======================================================
    # =======================================================

    proyectos_data = {
        'estados': [
            Proyecto.objects.filter(estado='APROBADO').count(),
            Proyecto.objects.filter(estado='ENVIADO').count(),
            Proyecto.objects.filter(estado='RECHAZADO').count(),
            Proyecto.objects.filter(estado='BORRADOR').count()
        ],
        'programado_mensual': [], 'realizado_mensual': []
    }
    for mes in range(1, 13):
        programado = AvanceMensual.objects.filter(mes=mes).aggregate(total=Count('cantidad_programada_mes'))['total'] or 0
        realizado = AvanceMensual.objects.filter(mes=mes).aggregate(total=Count('cantidad_realizada'))['total'] or 0
        proyectos_data['programado_mensual'].append(programado)
        proyectos_data['realizado_mensual'].append(realizado)
    
    total_unidades = unidades_con_datos.count() 
    total_proyectos = Proyecto.objects.count()
    proyectos_aprobados = Proyecto.objects.filter(estado='APROBADO').count()
    proyectos_pendientes = Proyecto.objects.filter(estado='ENVIADO').count()
    
    contexto = {
        'titulo': 'Estadísticas y Gráficos',
        'unidades_data': json.dumps(unidades_data),
        'proyectos_data': json.dumps(proyectos_data),
        'total_unidades': total_unidades,
        'total_proyectos': total_proyectos,
        'proyectos_aprobados': proyectos_aprobados,
        'proyectos_pendientes': proyectos_pendientes,
        'unidades_trimestrales_page': page_obj,
        'busqueda': busqueda,
        'active_tab': active_tab,
    }
    
    return render(request, 'administrador/estadisticas.html', contexto)

@admin_required
def lista_unidades(request):
    """Lista todas las unidades con buscador dinámico"""
    
    busqueda = request.GET.get('buscar', '')
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    
    if busqueda:
        unidades = unidades.filter(
            Q(unidad__nombre__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    unidades = unidades.annotate(
        total_proyectos=Count('proyectos'),
        count_proyectos_aprobados=Count('proyectos', filter=Q(proyectos__estado='APROBADO'))
    )
    
    unidades_con_rendimiento = []
    for unidad in unidades:
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento_promedio = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            unidad.rendimiento = round(cumplimiento_promedio, 2)
        else:
            unidad.rendimiento = 0
        unidades_con_rendimiento.append(unidad)
    
    contexto = {
        'titulo': 'Gestión de Unidades',
        'unidades': unidades_con_rendimiento,
        'busqueda': busqueda,
    }
    
    return render(request, 'administrador/lista_unidades.html', contexto)


@admin_required
def proyectos_unidad(request, unidad_id):
    """Lista todos los proyectos de una unidad específica"""
    from poa.models import ObjetivoEstrategico
    
    unidad = get_object_or_404(Usuario, id=unidad_id, rol='UNIDAD')
    proyectos = Proyecto.objects.filter(unidad=unidad).order_by('-anio', '-fecha_creacion')
    objetivos_estrategicos = ObjetivoEstrategico.objects.filter(activa=True)
    
    contexto = {
        'titulo': f'Proyectos de {unidad.unidad.nombre}',
        'unidad': unidad,
        'proyectos': proyectos,
        'objetivos_estrategicos': objetivos_estrategicos,
    }
    
    return render(request, 'administrador/proyectos_unidad.html', contexto)



@admin_required
def aprobar_proyecto(request, proyecto_id):
    """
    Lógica corregida: Aprueba el proyecto SIN eliminar los demás.
    Permite que una unidad tenga múltiples proyectos aprobados (ej. diferentes años o fondos).
    """
    if request.user.rol != 'ADMIN':
        return redirect('login:login')
        
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        # 1. Cambiar estado a Aprobado
        proyecto.estado = 'APROBADO'
        proyecto.aprobado_por = request.user
        proyecto.fecha_aprobacion = timezone.now()
        proyecto.motivo_rechazo = '' # Limpiar motivos de rechazo anteriores
        proyecto.save()
        
        # CORRECCIÓN: Se eliminó el bloque que borraba los demás proyectos de la unidad.
        # Antes: Proyecto.objects.filter(unidad=proyecto.unidad).exclude(id=proyecto.id).delete() -> ELIMINADO
        
        messages.success(request, f'El proyecto "{proyecto.nombre}" ha sido aprobado exitosamente.')
        return redirect('administrador:dashboard')
    
    return render(request, 'administrador/confirmar_aprobacion.html', {'proyecto': proyecto})


@admin_required
def rechazar_proyecto(request, proyecto_id):
    """Rechaza un proyecto enviado por una unidad"""
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', '')
        
        if not motivo:
            messages.error(request, 'Debes proporcionar un motivo de rechazo.')
            return redirect('administrador:proyectos_unidad', unidad_id=proyecto.unidad.id)
        
        proyecto.estado = 'RECHAZADO'
        proyecto.motivo_rechazo = motivo
        proyecto.save()
        
        # Registrar en auditoría
        AuditoriaLog.objects.create(
            usuario=request.user,
            accion='RECHAZO',
            tabla='Proyecto',
            registro_id=proyecto.id,
            datos_nuevos={'estado': 'RECHAZADO', 'motivo': motivo},
            ip=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'Proyecto "{proyecto.nombre}" rechazado.')
        return redirect('administrador:proyectos_unidad', unidad_id=proyecto.unidad.id)
    
    contexto = {
        'titulo': 'Rechazar Proyecto',
        'proyecto': proyecto,
    }
    
    return render(request, 'administrador/rechazar_proyecto.html', contexto)


@admin_required
def detalle_proyecto_admin(request, proyecto_id):
    """Muestra el detalle completo de un proyecto para el admin"""
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    metas = proyecto.metas.prefetch_related('actividades__avances', 'actividades__evidencias')
    
    # Total de actividades
    total_actividades = Actividad.objects.filter(meta__proyecto=proyecto).count()
    
    # Presupuesto total (suma de total_recursos de todas las actividades)
    presupuesto_total = Actividad.objects.filter(meta__proyecto=proyecto).aggregate(
        total=Sum('total_recursos')
    )['total'] or Decimal('0.00')
    
    # Total de evidencias
    total_evidencias = Evidencia.objects.filter(actividad__meta__proyecto=proyecto).count()
    
    for meta in metas:
        for actividad in meta.actividades.all():
            for avance in actividad.avances.all():
                # Contar evidencias del mes específico
                avance.evidencias_count = Evidencia.objects.filter(
                    actividad=actividad,
                    mes=avance.mes
                ).count()
    
    contexto = {
        'titulo': f'Detalle: {proyecto.nombre}',
        'proyecto': proyecto,
        'metas': metas,
        'total_actividades': total_actividades,
        'presupuesto_total': presupuesto_total,
        'total_evidencias': total_evidencias,
    }
    
    return render(request, 'administrador/detalle_proyecto.html', contexto)


@admin_required
def editar_proyecto_admin(request, proyecto_id):
    """Permite al admin editar un proyecto completo tipo wizard"""
    
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener el paso actual de la sesión
    paso_actual = request.session.get(f'paso_edicion_{proyecto_id}', 1)
    
    # Obtener metas y actividades
    metas = proyecto.metas.prefetch_related('actividades__avances').all()
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'siguiente':
            paso_actual = min(paso_actual + 1, 4)
            request.session[f'paso_edicion_{proyecto_id}'] = paso_actual
        elif accion == 'anterior':
            paso_actual = max(paso_actual - 1, 1)
            request.session[f'paso_edicion_{proyecto_id}'] = paso_actual
        elif accion == 'guardar_info':
            # Guardar información básica del proyecto
            formulario = FormularioProyecto(request.POST, instance=proyecto)
            if formulario.is_valid():
                formulario.save()
                messages.success(request, 'Información del proyecto actualizada.')
                paso_actual = 2
                request.session[f'paso_edicion_{proyecto_id}'] = paso_actual
        elif accion == 'agregar_meta':
            # Agregar nueva meta
            formulario_meta = FormularioMeta(request.POST)
            if formulario_meta.is_valid():
                meta = formulario_meta.save(commit=False)
                meta.proyecto = proyecto
                meta.save()
                messages.success(request, 'Meta agregada exitosamente.')
        elif accion == 'agregar_actividad':
            # Agregar nueva actividad
            meta_id = request.POST.get('meta_id')
            meta = get_object_or_404(MetaProyecto, id=meta_id, proyecto=proyecto)
            formulario_actividad = FormularioActividad(request.POST)
            if formulario_actividad.is_valid():
                actividad = formulario_actividad.save(commit=False)
                actividad.meta = meta
                actividad.save()
                messages.success(request, 'Actividad agregada exitosamente.')
        elif accion == 'guardar_programacion':
            # Guardar programación mensual
            actividad_id = request.POST.get('actividad_id')
            actividad = get_object_or_404(Actividad, id=actividad_id, meta__proyecto=proyecto)
            
            for mes in range(1, 13):
                cantidad = request.POST.get(f'mes_{mes}', 0)
                try:
                    cantidad = int(cantidad)
                except ValueError:
                    cantidad = 0
                
                avance, created = AvanceMensual.objects.get_or_create(
                    actividad=actividad,
                    mes=mes,
                    defaults={'cantidad_programada_mes': cantidad}
                )
                if not created:
                    avance.cantidad_programada_mes = cantidad
                    avance.save()
            
            messages.success(request, 'Programación mensual actualizada.')
        elif accion == 'finalizar':
            # Limpiar sesión y redirigir
            if f'paso_edicion_{proyecto_id}' in request.session:
                del request.session[f'paso_edicion_{proyecto_id}']
            
            AuditoriaLog.objects.create(
                usuario=request.user,
                accion='EDICION_COMPLETA_ADMIN',
                tabla='Proyecto',
                registro_id=proyecto.id,
                datos_nuevos={'mensaje': 'Edición completa del proyecto via wizard'},
                ip=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Proyecto actualizado exitosamente.')
            return redirect('administrador:detalle_proyecto_admin', proyecto_id=proyecto.id)
    
    formulario_proyecto = FormularioProyecto(instance=proyecto)
    formulario_meta = FormularioMeta()
    formulario_actividad = FormularioActividad()
    
    actividades = Actividad.objects.filter(meta__proyecto=proyecto).select_related('meta')
    
    programacion_mensual = {}
    for actividad in actividades:
        programacion_mensual[actividad.id] = {}
        for avance in actividad.avances.all():
            programacion_mensual[actividad.id][avance.mes] = avance.cantidad_programada_mes

    contexto = {
        'titulo': f'Editar: {proyecto.nombre}',
        'proyecto': proyecto,
        'paso_actual': paso_actual,
        'metas': metas,
        'actividades': actividades,
        'formulario_proyecto': formulario_proyecto,
        'formulario_meta': formulario_meta,
        'formulario_actividad': formulario_actividad,
        'meses': range(1, 13),
        'programacion_mensual': programacion_mensual,
    }
    
    return render(request, 'administrador/editar_proyecto.html', contexto)


@admin_required
def buscar_unidades_ajax(request):
    """Búsqueda AJAX de unidades para el buscador dinámico"""
    
    busqueda = request.GET.get('q', '')
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    
    if busqueda:
        unidades = unidades.filter(
            Q(unidad__nombre__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    unidades = unidades.annotate(
        total_proyectos=Count('proyectos'),
        count_proyectos_aprobados=Count('proyectos', filter=Q(proyectos__estado='APROBADO'))
    )[:10]
    
    resultados = []
    for unidad in unidades:
        resultados.append({
            'id': unidad.id,
            'nombre': unidad.unidad.nombre,
            'email': unidad.email,
            'total_proyectos': unidad.total_proyectos,
            'proyectos_aprobados': unidad.count_proyectos_aprobados,
        })
    
    return JsonResponse({'unidades': resultados})


@admin_required
def exportar_proyecto_detalle_pdf(request, proyecto_id):
    """Exporta el detalle COMPLETO de un proyecto a PDF - Usa función compartida"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    response = generar_pdf_proyecto_detalle(proyecto, request.user)
    
    # Registrar en auditoría
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_PDF',
        tabla='Proyecto',
        registro_id=proyecto.id,
        datos_nuevos={'tipo': 'PDF_DETALLADO', 'proyecto': proyecto.nombre},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@admin_required
def exportar_proyecto_detalle_excel(request, proyecto_id):
    """Exporta el detalle COMPLETO de un proyecto a Excel - Usa función compartida"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    response = generar_excel_proyecto_detalle(proyecto, request.user)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_EXCEL',
        tabla='Proyecto',
        registro_id=proyecto.id,
        datos_nuevos={'tipo': 'EXCEL_DETALLADO', 'proyecto': proyecto.nombre},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@admin_required
def exportar_unidades_pdf(request):
    """Exporta el reporte de todas las unidades con su cumplimiento a PDF - Usa función compartida"""
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    unidades = unidades.annotate(
        total_proyectos=Count('proyectos'),
        count_proyectos_aprobados=Count('proyectos', filter=Q(proyectos__estado='APROBADO'))
    )
    
    unidades_con_rendimiento = []
    for unidad in unidades:
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento_promedio = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            unidad.rendimiento = round(cumplimiento_promedio, 2)
        else:
            unidad.rendimiento = 0
        
        if unidad.rendimiento >= 80:
            unidad.categoria = 'Excelente'
        elif unidad.rendimiento >= 60:
            unidad.categoria = 'Bueno'
        elif unidad.rendimiento >= 40:
            unidad.categoria = 'Regular'
        else:
            unidad.categoria = 'Bajo'
        
        unidades_con_rendimiento.append(unidad)
    
    response = generar_pdf_unidades(unidades_con_rendimiento, request.user)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_PDF',
        tabla='Usuario',
        registro_id=0,
        datos_nuevos={'tipo': 'PDF_UNIDADES', 'total_unidades': len(unidades_con_rendimiento)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@admin_required
def exportar_unidades_excel(request):
    """Exporta el reporte de todas las unidades con su cumplimiento a Excel - Usa función compartida"""
    
    unidades = Usuario.objects.filter(rol='UNIDAD').select_related('unidad')
    unidades = unidades.annotate(
        total_proyectos=Count('proyectos'),
        count_proyectos_aprobados=Count('proyectos', filter=Q(proyectos__estado='APROBADO'))
    )
    
    unidades_con_rendimiento = []
    for unidad in unidades:
        poa_aprobado = Proyecto.objects.filter(unidad=unidad, estado='APROBADO').first()
        if poa_aprobado:
            avances = AvanceMensual.objects.filter(actividad__meta__proyecto=poa_aprobado)
            cumplimiento_promedio = avances.aggregate(Avg('cumplimiento'))['cumplimiento__avg'] or 0
            unidad.rendimiento = round(cumplimiento_promedio, 2)
        else:
            unidad.rendimiento = 0
        
        if unidad.rendimiento >= 80:
            unidad.categoria = 'Excelente'
            unidad.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        elif unidad.rendimiento >= 60:
            unidad.categoria = 'Bueno'
            unidad.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
        elif unidad.rendimiento >= 40:
            unidad.categoria = 'Regular'
            unidad.fill = PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid")
        else:
            unidad.categoria = 'Bajo'
            unidad.fill = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")
        
        unidades_con_rendimiento.append(unidad)
    
    response = generar_excel_unidades(unidades_con_rendimiento, request.user)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_EXCEL',
        tabla='Usuario',
        registro_id=0,
        datos_nuevos={'tipo': 'EXCEL_UNIDADES', 'total_unidades': len(unidades_con_rendimiento)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response



@admin_required
def exportar_reporte_trimestral_pdf(request):
    """Exporta el reporte trimestral (filtrado) a PDF"""
    
    busqueda = request.GET.get('buscar', '')
    
    unidades_trimestrales = _get_datos_trimestrales(busqueda)
    
    response = generar_pdf_reporte_trimestral(unidades_trimestrales, request.user, busqueda)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_PDF',
        tabla='Reporte',
        registro_id=0,
        datos_nuevos={'tipo': 'PDF_REPORTE_TRIMESTRAL', 'filtro': busqueda, 'total': len(unidades_trimestrales)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response

@admin_required
def exportar_reporte_trimestral_excel(request):
    """Exporta el reporte trimestral (filtrado) a Excel"""
    
    busqueda = request.GET.get('buscar', '')
    
    unidades_trimestrales = _get_datos_trimestrales(busqueda)
    
    response = generar_excel_reporte_trimestral(unidades_trimestrales, request.user, busqueda)
    
    AuditoriaLog.objects.create(
        usuario=request.user,
        accion='EXPORTACION_EXCEL',
        tabla='Reporte',
        registro_id=0,
        datos_nuevos={'tipo': 'EXCEL_REPORTE_TRIMESTRAL', 'filtro': busqueda, 'total': len(unidades_trimestrales)},
        ip=request.META.get('REMOTE_ADDR')
    )
    
    return response


@admin_required
def lista_metas_predeterminadas(request):
    """Lista las metas predeterminadas"""
    from poa.models import MetaPredeterminada
    metas = MetaPredeterminada.objects.all()
    return render(request, 'administrador/lista_metas.html', {'metas': metas})

@admin_required
def crear_meta_predeterminada(request):
    """Crea una nueva meta predeterminada"""
    from poa.models import MetaPredeterminada
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        activa = request.POST.get('activa') == 'on'
        
        if nombre:
            MetaPredeterminada.objects.create(nombre=nombre, activa=activa)
            messages.success(request, 'Meta predeterminada creada exitosamente.')
            return redirect('administrador:lista_metas_predeterminadas')
            
    return render(request, 'administrador/form_meta.html')

@admin_required
def editar_meta_predeterminada(request, meta_id):
    """Edita una meta predeterminada existente"""
    from poa.models import MetaPredeterminada
    meta = get_object_or_404(MetaPredeterminada, id=meta_id)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        activa = request.POST.get('activa') == 'on'
        
        if nombre:
            meta.nombre = nombre
            meta.activa = activa
            meta.save()
            messages.success(request, 'Meta predeterminada actualizada exitosamente.')
            return redirect('administrador:lista_metas_predeterminadas')
            
    return render(request, 'administrador/form_meta.html', {'meta': meta})

@admin_required
def eliminar_meta_predeterminada(request, meta_id):
    """Elimina una meta predeterminada"""
    from poa.models import MetaPredeterminada
    meta = get_object_or_404(MetaPredeterminada, id=meta_id)
    meta.delete()
    messages.success(request, 'Meta predeterminada eliminada exitosamente.')
    return redirect('administrador:lista_metas_predeterminadas')

# --- Vistas para Objetivos Estratégicos ---

@admin_required
def lista_objetivos_estrategicos(request):
    """Lista los objetivos estratégicos"""
    from poa.models import ObjetivoEstrategico
    objetivos = ObjetivoEstrategico.objects.all()
    return render(request, 'administrador/lista_objetivos.html', {'objetivos': objetivos})

@admin_required
def crear_objetivo_estrategico(request):
    """Crea un nuevo objetivo estratégico"""
    from poa.models import ObjetivoEstrategico
    
    if request.method == 'POST':
        descripcion = request.POST.get('descripcion')
        activa = request.POST.get('activa') == 'on'
        
        if descripcion:
            ObjetivoEstrategico.objects.create(
                descripcion=descripcion,
                activa=activa
            )
            messages.success(request, 'Objetivo estratégico creado correctamente.')
            return redirect('administrador:lista_objetivos_estrategicos')
        else:
            messages.error(request, 'La descripción es obligatoria.')
            
    return render(request, 'administrador/form_objetivo.html')

@admin_required
def editar_objetivo_estrategico(request, objetivo_id):
    """Edita un objetivo estratégico existente"""
    from poa.models import ObjetivoEstrategico
    objetivo = get_object_or_404(ObjetivoEstrategico, id=objetivo_id)
    
    if request.method == 'POST':
        descripcion = request.POST.get('descripcion')
        activa = request.POST.get('activa') == 'on'
        
        if descripcion:
            objetivo.descripcion = descripcion
            objetivo.activa = activa
            objetivo.save()
            messages.success(request, 'Objetivo estratégico actualizado correctamente.')
            return redirect('administrador:lista_objetivos_estrategicos')
        else:
            messages.error(request, 'La descripción es obligatoria.')
            
    return render(request, 'administrador/form_objetivo.html', {'objetivo': objetivo})

@admin_required
def eliminar_objetivo_estrategico(request, objetivo_id):
    """Elimina un objetivo estratégico"""
    from poa.models import ObjetivoEstrategico
    objetivo = get_object_or_404(ObjetivoEstrategico, id=objetivo_id)
    
    if request.method == 'POST':
        objetivo.delete()
        messages.success(request, 'Objetivo estratégico eliminado correctamente.')
        return redirect('administrador:lista_objetivos_estrategicos')
        
    return render(request, 'administrador/confirmar_eliminar.html', {
        'objeto': objetivo,
        'tipo': 'Objetivo Estratégico',
        'url_cancelar': 'administrador:lista_objetivos_estrategicos'
    })

@admin_required
def exportar_proyectos_unidad(request, unidad_id):
    """Exporta los proyectos de una unidad a Excel en formato POA"""
    from poa.models import ObjetivoEstrategico
    from .excel_export import generar_poa_excel
    from io import BytesIO
    
    unidad_usuario = get_object_or_404(Usuario, id=unidad_id, rol='UNIDAD')
    unidad = unidad_usuario.unidad
    
    # Obtener proyectos aprobados del año actual (excluyendo no planificados de la lista principal)
    # La función generar_poa_excel se encargará de buscar el proyecto no planificado por separado
    proyectos = Proyecto.objects.filter(
        unidad=unidad_usuario, 
        estado='APROBADO',
        es_no_planificado=False
    ).order_by('-anio', 'fecha_creacion')
    
    # Obtener el objetivo estratégico seleccionado
    objetivo_id = request.GET.get('objetivo_estrategico')
    objetivo_estrategico = None
    if objetivo_id:
        objetivo_estrategico = get_object_or_404(ObjetivoEstrategico, id=objetivo_id)
    
    # Obtener solo los proyectos APROBADOS de la unidad (excluyendo no planificados)
    proyectos = Proyecto.objects.filter(
        unidad=unidad_usuario,
        estado='APROBADO',
        es_no_planificado=False
    ).prefetch_related('metas__actividades')
    
    # Generar el Excel
    wb = generar_poa_excel(unidad, proyectos, objetivo_estrategico)
    
    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'POA_{unidad.nombre.replace(" ", "_")}_{datetime.now().year}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar el workbook en la respuesta
    wb.save(response)
    
    return response